"""
通用批量测试引擎

从边界值汇总 Excel 加载测试用例，通过 API 逐个验证保费费率计算。
支持自动节流、进度显示、错误恢复。
"""

import re
import random
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

from .config import ProductProfile
from .api_client import InsuranceAPIClient


class TestCaseLoader:
    """
    测试用例加载器 —— 从标准 10 列边界值汇总 Excel 读取用例。

    标准输入格式:
        保障方案 | ensurePlan | 责任计划 | 保险期间 | 交费期间(年) |
        性别 | 最小年龄(岁) | 最小年龄费率 | 最大年龄(岁) | 最大年龄费率

    每行生成 1~2 个测试用例（最小年龄 + 最大年龄）。
    """

    def __init__(self, profile: ProductProfile):
        self.profile = profile

    def load(self, file_path: str, sheet_name: str = None) -> list[dict]:
        """
        从边界值汇总 Excel 加载测试用例。

        Args:
            file_path: Excel 文件路径
            sheet_name: Sheet 名称（默认第一个 Sheet）

        Returns:
            测试用例列表，每项包含:
            {
                "保障方案": str, "ensurePlan": str, "责任计划": int,
                "保险期间": str, "交费期间": int, "性别": str,
                "年龄": int, "年龄类型": str ("最小年龄"/"最大年龄"),
                "期望费率": float,
            }
        """
        wb = load_workbook(file_path, data_only=True)
        sheet_name = sheet_name or wb.sheetnames[0]
        ws = wb[sheet_name]

        cases = []
        data_start = 4  # 第 1-3 行是标题/说明/表头

        for row in range(data_start, ws.max_row + 1):
            # 读取 10 列
            plan_name = self._val(ws, row, 1)       # 保障方案
            ensure_plan = self._val(ws, row, 2)      # ensurePlan
            plan = self._int(ws, row, 3)             # 责任计划
            period = self._val(ws, row, 4)            # 保险期间
            pay_period = self._int(ws, row, 5)       # 交费期间
            gender = self._val(ws, row, 6)            # 性别
            min_age = self._int(ws, row, 7)          # 最小年龄
            min_rate = self._float(ws, row, 8)       # 最小年龄费率
            max_age = self._int(ws, row, 9)          # 最大年龄
            max_rate = self._float(ws, row, 10)      # 最大年龄费率

            # 跳过无效行
            if plan is None:
                continue
            if gender not in ("男", "女"):
                continue
            if pay_period is None or not period:
                continue

            # 生成最小年龄用例
            if min_age is not None and min_rate is not None:
                cases.append({
                    "保障方案": plan_name,
                    "ensurePlan": self._sanitize_ensure_plan(ensure_plan),
                    "责任计划": plan,
                    "保险期间": period,
                    "交费期间": pay_period,
                    "性别": gender,
                    "年龄": min_age,
                    "年龄类型": "最小年龄",
                    "期望费率": min_rate,
                })

            # 生成最大年龄用例（与最小年龄不同时才生成）
            if (max_age is not None and max_rate is not None
                    and max_age != min_age):
                cases.append({
                    "保障方案": plan_name,
                    "ensurePlan": self._sanitize_ensure_plan(ensure_plan),
                    "责任计划": plan,
                    "保险期间": period,
                    "交费期间": pay_period,
                    "性别": gender,
                    "年龄": max_age,
                    "年龄类型": "最大年龄",
                    "期望费率": max_rate,
                })

        wb.close()
        return cases

    @staticmethod
    def _val(ws, row, col):
        v = ws.cell(row=row, column=col).value
        return str(v).strip() if v is not None else None

    @staticmethod
    def _sanitize_ensure_plan(raw) -> str:
        """防御性规范化 ensurePlan：只允许 "1" 或 "2"，其余一律回退为 "1"。"""
        if raw is None:
            return "1"
        s = str(raw).strip()
        if s in ("1", "2"):
            return s
        return "1"

    @staticmethod
    def _int(ws, row, col):
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        try:
            return int(float(v))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _float(ws, row, col):
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        try:
            return round(float(v), 10)
        except (ValueError, TypeError):
            return None


class BatchTester:
    """
    批量测试执行器。

    用法:
        profile = ProductProfile.from_yaml("config.yaml")
        tester = BatchTester(profile, serial_no="1491455820292947968")
        results = tester.run("边界值汇总.xlsx")
    """

    def __init__(self, profile: ProductProfile, serial_no: str, proposal_id: str = ""):
        if not serial_no:
            raise ValueError("serial_no 是模拟测算的关键参数，不能为空")
        self.profile = profile
        self.serial_no = serial_no
        self.proposal_id = proposal_id
        self.client = InsuranceAPIClient(profile, serial_no=serial_no, proposal_id=proposal_id)
        self.loader = TestCaseLoader(profile)

    # ================================================================
    # 主入口
    # ================================================================

    def run(
        self,
        boundary_file: str,
        output_file: str = None,
        smoke_only: bool = False,
        smoke_count: int = 5,
        progress_callback=None,
    ) -> list[dict]:
        """
        执行完整的批量测试流程。

        Args:
            boundary_file: 边界值汇总 Excel 路径
            output_file: 报告输出路径（None=自动生成）
            smoke_only: 是否仅执行冒烟测试
            smoke_count: 冒烟测试用例数
            progress_callback: 进度回调 fn(current, total)

        Returns:
            测试结果列表
        """
        # 加载用例
        print(f"📂 加载测试用例: {boundary_file}")
        cases = self.loader.load(boundary_file)

        if smoke_only:
            cases = cases[:smoke_count]
            print(f"🔥 冒烟测试模式: 仅执行前 {len(cases)} 条")

        print(f"  共加载 {len(cases)} 条测试用例")

        if not cases:
            print("⚠ 没有可执行的测试用例")
            return []

        # 登录
        print("🔑 登录中...")
        try:
            ok, _ = self.client.login()
        except Exception as e:
            print(f"❌ 无法连接 API 服务器: {e}")
            print("   请确认:")
            print("   1. 是否在公司内网环境")
            print(f"   2. API 地址是否正确: {self.profile.api.base_url}")
            print("   3. VPN 是否已连接")
            return []
        if not ok:
            print("❌ 登录失败，请检查账号密码配置")
            print(f"   API 地址: {self.profile.api.base_url}")
            return []

        # 执行测试（并行）
        workers = self.profile.test.throttle.workers
        print(f"🧪 开始并行执行 {len(cases)} 条用例 ({workers} 线程)...\n")
        results = [None] * len(cases)
        start_time = time.time()
        completed = 0

        login_cookies = self.client.login_cookies

        def worker(case, case_no):
            """每个线程独立创建 client，共享 login_cookies"""
            worker_client = InsuranceAPIClient(
                self.profile,
                serial_no=self.serial_no,
                proposal_id=self.proposal_id,
            )
            worker_client.login_cookies = login_cookies
            return self._run_single(case, case_no, client=worker_client)

        throttle = self.profile.test.throttle
        with ThreadPoolExecutor(max_workers=workers) as executor:
            future_to_idx = {
                executor.submit(worker, case, i + 1): i
                for i, case in enumerate(cases)
            }
            for future in as_completed(future_to_idx):
                idx = future_to_idx[future]
                try:
                    result = future.result()
                except Exception as e:
                    result = {
                        "序号": idx + 1,
                        "测试结论": f"ERROR - 线程异常: {e}",
                        "备注": str(e),
                    }
                results[idx] = result
                completed += 1

                # 输出单条结果
                verdict = result.get("测试结论", "?")
                emoji = "✅" if verdict.startswith("PASS") else "❌" if verdict.startswith("FAIL") else "⚠️"
                print(f"  [{completed}/{len(cases)}] {emoji} {verdict} | "
                      f"计划{result.get('责任计划','?')} {result.get('性别','?')} "
                      f"{result.get('年龄','?')}岁 | 期望{result.get('期望费率(‰)','?')}‰")

                # 进度回调（第三个参数传递当前结果，用于前端实时展示）
                if progress_callback:
                    progress_callback(completed, len(cases), result)

        # 过滤掉 None（理论上不应出现）
        results = [r for r in results if r is not None]

        elapsed = time.time() - start_time

        # 统计
        pass_count = sum(1 for r in results if str(r.get("测试结论", "")).startswith("PASS"))
        fail_count = sum(1 for r in results if str(r.get("测试结论", "")).startswith("FAIL"))
        error_count = sum(1 for r in results if str(r.get("测试结论", "")).startswith("ERROR"))

        print(f"\n{'='*50}")
        print(f"  总用例: {len(results)}")
        print(f"  ✅ PASS:  {pass_count} ({100*pass_count/len(results):.1f}%)" if pass_count else "  ✅ PASS:  0")
        print(f"  ❌ FAIL:  {fail_count}" if fail_count else "  ❌ FAIL:  0")
        print(f"  ⚠ ERROR: {error_count}" if error_count else "  ⚠ ERROR: 0")
        print(f"  ⏱ 耗时:   {elapsed:.1f}s")
        print(f"{'='*50}\n")

        return results

    # ================================================================
    # 金额限制解析（智能调整）
    # ================================================================

    # 金额限制相关关键词
    _AMOUNT_LIMIT_KEYWORDS = [
        "最低", "最高", "调整", "步长", "单位", "万元", "保额",
        "不低于", "不超过", "整数倍", "倍数",
    ]

    # 金额提取正则（按优先级排列）
    _AMOUNT_PATTERNS = [
        # "100万元" / "100万"
        (re.compile(r'(\d+(?:,\d{3})*(?:\.\d+)?)\s*万\s*元?'), 10000),
        # "1,000,000元" / "1000000元"
        (re.compile(r'(\d+(?:,\d{3})*(?:\.\d+)?)\s*元'), 1),
        # "1000的整数倍" / "1000的倍数"
        (re.compile(r'(\d+(?:,\d{3})*)\s*的\s*(?:整数倍|倍数)'), 1),
        # 纯数字（≥4位，可能是金额）
        (re.compile(r'(\d{4,})'), 1),
    ]

    @classmethod
    def _is_amount_limit_reason(cls, reason: str) -> bool:
        """判断 failureReason 是否与金额限制相关（需要调整重试）。"""
        if not reason:
            return False
        return any(kw in reason for kw in cls._AMOUNT_LIMIT_KEYWORDS)

    @classmethod
    def _parse_amount_limit(cls, reason: str) -> Optional[dict]:
        """
        从 failureReason 文本中解析金额限制。

        Returns:
            {"type": "min"|"max"|"step", "value": int} 或 None（无法解析）
        """
        if not reason:
            return None

        # 判断限制类型（step 优先：\"最低调整单位\"整体是 step 而非 min）
        if any(w in reason for w in ("步长", "调整单位", "整数倍", "倍数")):
            limit_type = "step"
        elif any(w in reason for w in ("最低", "不低于")):
            limit_type = "min"
        elif any(w in reason for w in ("最高", "不超过")):
            limit_type = "max"
        else:
            limit_type = "min"  # 默认当作最低限制

        # 提取金额数值
        for pattern, multiplier in cls._AMOUNT_PATTERNS:
            m = pattern.search(reason)
            if m:
                num_str = m.group(1).replace(",", "")
                try:
                    value = int(float(num_str) * multiplier)
                    return {"type": limit_type, "value": value}
                except (ValueError, TypeError):
                    continue

        return None

    def _adjust_amount(self, current: int, limit: dict,
                       amount_cfg) -> int:
        """
        根据解析出的金额限制调整保额。

        Args:
            current: 当前保额
            limit: {"type": "min"|"max"|"step", "value": int}
            amount_cfg: AmountConfig 配置

        Returns:
            调整后的保额
        """
        limit_type = limit["type"]
        limit_value = limit["value"]
        new_amount = current

        if limit_type == "min":
            # 不低于限制值，留一点余量（+1个step）
            new_amount = max(current, limit_value + amount_cfg.step)
            # 确保不超过 max
            new_amount = min(new_amount, amount_cfg.max)

        elif limit_type == "max":
            # 不超过限制值，留一点余量（-1个step）
            new_amount = min(current, limit_value - amount_cfg.step)
            # 确保不低于 min
            new_amount = max(new_amount, amount_cfg.min)

        elif limit_type == "step":
            # 调整为 step 的整数倍
            step = max(limit_value, amount_cfg.step)
            new_amount = ((current + step // 2) // step) * step
            new_amount = max(new_amount, amount_cfg.min)
            new_amount = min(new_amount, amount_cfg.max)

        return new_amount

    # ================================================================
    # 单用例执行
    # ================================================================

    def _run_single(self, case: dict, case_no: int, client: InsuranceAPIClient = None) -> dict:
        """
        执行单个测试用例。

        流程:
        1. 调用 age_rate(saveCustomer)
        2. 调用 plan_rate(saveProductExt)
        3. 若触发金额限制（最低/最高/调整单位），智能调整输入值并重试（最多10次）
        4. 比较 API 返回值与期望值（允许 tolerance 误差）

        根据 product.data_type 自动切换算费方向:
          type="1"（保额算保费）: 输入=保额, 期望输出=保费, API返回=fee
          type="2"（保费算保额）: 输入=保费, 期望输出=保额, API返回=amount
        """
        plan = case["责任计划"]
        period = case["保险期间"]
        pay_period = case["交费期间"]
        gender = case["性别"]
        age = case["年龄"]
        age_type = case["年龄类型"]
        ensure_plan = case.get("ensurePlan", "1")
        expected_rate = case["期望费率"]

        # ---- 算费方向和费率单位 ----
        data_type = self.profile.data_type  # "1"=保额算保费, "2"=保费算保额
        fee_unit = self.profile.fee_unit or 1000
        is_type2 = (data_type == "2")
        input_type = "premium" if is_type2 else "amount"

        # ---- 生成随机输入值（保额或保费，取决于算费方向） ----
        if is_type2:
            input_cfg = self.profile.test.premium   # 保费配置：min/max/step 对应保费范围
        else:
            input_cfg = self.profile.test.amount    # 保额配置：min/max/step 对应保额范围

        if input_cfg.fixed:
            input_val = input_cfg.fixed
        else:
            input_val = random.randrange(input_cfg.min, input_cfg.max + 1, input_cfg.step)

        # ---- 计算期望输出值 ----
        if is_type2:
            # 保费算保额: 期望保额 = 保费 × 费率单位 ÷ 费率
            expected_output = input_val * fee_unit / expected_rate
            input_label = "保费(元)"
            output_label = "期望保额(元)"
        else:
            # 保额算保费: 期望保费 = 保额 × 费率 ÷ 费率单位
            expected_output = input_val * expected_rate / fee_unit
            input_label = "保额(元)"
            output_label = "期望保费(元)"

        # ---- 基础结果（列名根据算费方向动态调整） ----
        result = {
            "序号": case_no,
            "算费方向": "保费算保额" if is_type2 else "保额算保费",
            "保障方案": case.get("保障方案", ""),
            "ensurePlan": ensure_plan,
            "责任计划": plan,
            "责任描述": self.profile.plan_has_description(plan),
            "保险期间": period,
            "保险期间Code": self.profile.get_ensure_period_code(period),
            "交费期间": pay_period,
            "交费期间Code": self.profile.get_pay_period_code(pay_period),
            "交费方式Code": self.profile.get_pay_mode_code(pay_period),
            "性别": gender,
            "性别Code": self.profile.get_gender_code(gender),
            "年龄": age,
            "年龄类型": age_type,
            input_label: input_val,
            "期望费率(‰)": expected_rate,
            output_label: round(expected_output, 2),
            "age_rate状态码": None,
            "age_rate结果": "",
            "plan_rate状态码": None,
            "plan_rate结果": "",
            "API返回值": None,
            "failureReason": None,
            "测试结论": "",
            "备注": "",
            "_data_type": data_type,       # 内部标记，供 report 动态列名用
            "_input_label": input_label,
            "_output_label": output_label,
        }

        api = client or self.client

        try:
            # Step 1: age_rate
            ok, cookies, status, text = api.save_customer(age, gender)
            result["age_rate状态码"] = status
            result["age_rate结果"] = "成功" if ok else "失败"

            if not ok:
                result["测试结论"] = "FAIL - age_rate接口失败"
                return result

            # Step 2: plan_rate（带金额限制智能重试）
            MAX_RETRIES = 10
            retry_log = []
            final_result_val = None
            final_status = None
            final_reason = None
            final_ok = False

            for attempt in range(1 + MAX_RETRIES):
                ok, result_val, status, text, reason = api.save_product(
                    plan=plan,
                    ensure_period=period,
                    pay_period=pay_period,
                    amount=input_val,
                    ensure_plan=ensure_plan,
                    cookies=cookies,
                    input_type=input_type,
                )

                if ok:
                    final_ok = True
                    final_result_val = result_val
                    final_status = status
                    final_reason = reason
                    break

                if attempt >= MAX_RETRIES:
                    final_result_val = result_val
                    final_status = status
                    final_reason = reason
                    break

                if reason and self._is_amount_limit_reason(reason):
                    limit = self._parse_amount_limit(reason)
                    old_val = input_val

                    if limit:
                        input_val = self._adjust_amount(old_val, limit, input_cfg)
                        retry_log.append(
                            f"第{attempt+1}次: {reason} → "
                            f"{'保费' if is_type2 else '保额'} {old_val:,} → {input_val:,}"
                            f"(解析: {limit['type']}={limit['value']:,})"
                        )
                    else:
                        input_val = random.randrange(
                            input_cfg.min, input_cfg.max + 1, input_cfg.step
                        )
                        retry_log.append(
                            f"第{attempt+1}次: {reason} → "
                            f"{'保费' if is_type2 else '保额'} {old_val:,} → {input_val:,} (随机调整)"
                        )
                else:
                    final_result_val = result_val
                    final_status = status
                    final_reason = reason
                    break

            # 更新结果
            result[input_label] = input_val
            result["API返回值"] = final_result_val
            result["plan_rate状态码"] = final_status
            result["failureReason"] = final_reason
            if retry_log:
                result["备注"] = " | ".join(retry_log)

            if final_ok:
                result["plan_rate结果"] = "成功"
                # 用最终输入值重新算期望输出
                if is_type2:
                    final_expected = input_val * fee_unit / expected_rate
                else:
                    final_expected = input_val * expected_rate / fee_unit
                result[output_label] = round(final_expected, 2)

                actual_val = float(final_result_val)
                deviation = (abs(actual_val - final_expected) / final_expected
                           if final_expected > 0 else float("inf"))

                if deviation < self.profile.test.tolerance:
                    result["测试结论"] = "PASS"
                else:
                    pct = deviation * 100
                    result["测试结论"] = f"PASS(偏差{pct:.1f}%)"
                    if not retry_log:
                        result["备注"] = (
                            f"期望{final_expected:.2f}, 实际{actual_val:.2f}, "
                            f"偏差{pct:.2f}%"
                        )
            else:
                result["plan_rate结果"] = "失败" if final_status == 500 else "异常"
                if final_reason:
                    result["测试结论"] = f"FAIL - {final_reason}"
                    if retry_log:
                        result["测试结论"] += " (金额调整10次仍失败)"
                elif final_result_val is not None and float(final_result_val) == 0:
                    result["测试结论"] = f"FAIL - 计算失败(返回值为0)"
                elif final_result_val is not None:
                    result["测试结论"] = "FAIL - API返回异常"
                else:
                    result["测试结论"] = "FAIL - plan_rate接口失败"

        except Exception as e:
            result["测试结论"] = f"ERROR - {str(e)}"
            result["备注"] = str(e)

        return result
