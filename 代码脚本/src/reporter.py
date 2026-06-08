"""
通用 Excel 报告生成器

将批量测试结果生成格式化的 Excel 报告，包含：
- Sheet 1: 测试结果明细（24 列）
- Sheet 2: 汇总统计（多维度交叉统计）

颜色标记: PASS→绿, FAIL→红, ERROR→黄
"""

from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


class ReportGenerator:
    """
    Excel 测试报告生成器。

    用法:
        gen = ReportGenerator(profile)
        gen.generate(results, "API测试报告.xlsx")
    """

    # 24 列表头
    DETAIL_HEADERS = [
        "序号", "保障方案", "ensurePlan", "责任计划", "责任描述",
        "保险期间", "保险期间Code", "交费期间(年)", "交费期间Code", "交费方式Code",
        "性别", "性别Code", "年龄(岁)", "年龄类型", "保额(元)",
        "期望费率(‰)", "期望保费(元)",
        "age_rate\n状态码", "age_rate\n结果",
        "plan_rate\n状态码", "plan_rate\n结果",
        "API返回fee", "failureReason", "测试结论",
    ]

    COL_WIDTHS = [
        6, 10, 10, 8, 38, 8, 12, 10, 12, 10,
        6, 8, 8, 10, 14, 12, 14, 10, 10, 10,
        10, 14, 30, 18,
    ]

    # 条件颜色
    COLOR_PASS = "E2EFDA"    # 浅绿
    COLOR_FAIL = "FCE4D6"    # 浅橙/红
    COLOR_ERROR = "FFF2CC"   # 浅黄
    COLOR_ALT1 = "F2F7FB"    # 交替色 1 浅蓝
    COLOR_ALT2 = "FFFFFF"    # 交替色 2 白色

    def __init__(self, product_name: str = ""):
        self.product_name = product_name

    # ================================================================
    # 主入口
    # ================================================================

    def generate(
        self,
        results: list[dict],
        output_path: str,
        detail_sheet_name: str = "测试结果明细",
        summary_sheet_name: str = "汇总统计",
    ) -> str:
        """
        生成 Excel 测试报告。

        Args:
            results: 测试结果列表
            output_path: 输出文件路径
            detail_sheet_name: 明细 Sheet 名称
            summary_sheet_name: 汇总 Sheet 名称

        Returns:
            输出文件路径
        """
        wb = Workbook()

        # Sheet 1: 明细
        self._write_detail_sheet(wb, results, detail_sheet_name)

        # Sheet 2: 汇总
        self._write_summary_sheet(wb, results, summary_sheet_name)

        # 删除默认 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_path)
        print(f"✅ 测试报告已生成: {output_path}")
        return output_path

    # ================================================================
    # Sheet 1: 测试结果明细
    # ================================================================

    def _write_detail_sheet(self, wb: Workbook, results: list[dict], sheet_name: str):
        ws = wb.active
        ws.title = sheet_name

        n_total = len(results)
        n_pass = sum(1 for r in results if str(r.get("测试结论", "")).startswith("PASS"))
        n_fail = sum(1 for r in results if str(r.get("测试结论", "")).startswith("FAIL"))
        n_error = sum(1 for r in results if str(r.get("测试结论", "")).startswith("ERROR"))
        pass_rate = (n_pass / n_total * 100) if n_total > 0 else 0

        # ---- 第 1 行：标题 ----
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        title = f"{self.product_name} — API 测试报告  生成时间：{ts}" if self.product_name else f"API 测试报告  生成时间：{ts}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.DETAIL_HEADERS))
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # ---- 第 2 行：统计摘要 ----
        summary = (
            f"总用例: {n_total} | ✅ PASS: {n_pass} ({pass_rate:.1f}%) | "
            f"❌ FAIL: {n_fail} | ⚠ ERROR: {n_error}"
        )
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(self.DETAIL_HEADERS))
        c = ws.cell(row=2, column=1, value=summary)
        c.font = Font(name="微软雅黑", size=10, bold=True, color="333333")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 24

        # ---- 第 3 行：表头 ----
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")

        for col_idx, header in enumerate(self.DETAIL_HEADERS, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 36

        # ---- 数据行 ----
        thin_border = Border(
            left=Side(style="thin", color="E0E0E0"),
            right=Side(style="thin", color="E0E0E0"),
            top=Side(style="thin", color="E0E0E0"),
            bottom=Side(style="thin", color="E0E0E0"),
        )
        data_font = Font(name="微软雅黑", size=9)
        data_align = Alignment(horizontal="center", vertical="center")

        for i, result in enumerate(results):
            row_num = 4 + i
            verdict = str(result.get("测试结论", ""))

            # 选择行颜色
            if verdict.startswith("PASS"):
                row_color = self.COLOR_PASS
            elif verdict.startswith("FAIL"):
                row_color = self.COLOR_FAIL
            elif verdict.startswith("ERROR"):
                row_color = self.COLOR_ERROR
            else:
                row_color = self.COLOR_ALT1 if i % 2 == 0 else self.COLOR_ALT2

            row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")

            # 按顺序填充 24 列
            values = [
                result.get("序号", i + 1),
                result.get("保障方案", ""),
                result.get("ensurePlan", ""),
                result.get("责任计划", ""),
                result.get("责任描述", ""),
                result.get("保险期间", ""),
                result.get("保险期间Code", ""),
                result.get("交费期间", ""),
                result.get("交费期间Code", ""),
                result.get("交费方式Code", ""),
                result.get("性别", ""),
                result.get("性别Code", ""),
                result.get("年龄", ""),
                result.get("年龄类型", ""),
                result.get("保额(元)", ""),
                result.get("期望费率(‰)", ""),
                result.get("期望保费(元)", ""),
                result.get("age_rate状态码", ""),
                result.get("age_rate结果", ""),
                result.get("plan_rate状态码", ""),
                result.get("plan_rate结果", ""),
                result.get("API返回fee", ""),
                result.get("failureReason", ""),
                verdict,
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = data_font
                cell.fill = row_fill
                cell.alignment = data_align
                cell.border = thin_border

        # ---- 列宽 ----
        for col_idx, width in enumerate(self.COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ---- 冻结 ----
        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(self.DETAIL_HEADERS))}{3 + n_total}"

    # ================================================================
    # Sheet 2: 汇总统计
    # ================================================================

    def _write_summary_sheet(self, wb: Workbook, results: list[dict], sheet_name: str):
        ws = wb.create_sheet(sheet_name)

        # ---- 标题 ----
        ws.merge_cells("A1:F1")
        c = ws.cell(row=1, column=1, value="测试结果汇总统计")
        c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # 统计维度: 保障方案, 责任计划, 交费期间, 性别, 年龄类型
        dims = ["保障方案", "责任计划", "交费期间", "性别", "年龄类型"]
        current_row = 3

        for dim in dims:
            # 分组统计
            groups = defaultdict(lambda: {"total": 0, "pass": 0, "fail_error": 0})
            for r in results:
                key = str(r.get(dim, "未知"))
                groups[key]["total"] += 1
                verdict = str(r.get("测试结论", ""))
                if verdict.startswith("PASS"):
                    groups[key]["pass"] += 1
                else:
                    groups[key]["fail_error"] += 1

            # 写子标题
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=5)
            dim_label = {
                "保障方案": "按保障方案统计",
                "责任计划": "按责任计划统计",
                "交费期间": "按交费期间统计",
                "性别": "按性别统计",
                "年龄类型": "按年龄类型统计",
            }.get(dim, dim)
            c = ws.cell(row=current_row, column=1, value=f"📊 {dim_label}")
            c.font = Font(name="微软雅黑", size=11, bold=True, color="2E75B6")
            ws.row_dimensions[current_row].height = 24
            current_row += 1

            # 子表头
            sub_headers = [dim, "总用例", "PASS", "FAIL/ERROR", "通过率"]
            hdr_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            for col_idx, h in enumerate(sub_headers, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center")
            current_row += 1

            # 数据行
            sorted_keys = sorted(groups.keys())
            for i, key in enumerate(sorted_keys):
                g = groups[key]
                pass_rate = (g["pass"] / g["total"] * 100) if g["total"] > 0 else 0

                vals = [key, g["total"], g["pass"], g["fail_error"], f"{pass_rate:.1f}%"]
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    cell.font = Font(name="微软雅黑", size=9)
                    cell.alignment = Alignment(horizontal="center")

                    # 全部通过则绿色
                    if g["fail_error"] == 0 and col_idx >= 2:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    elif g["fail_error"] > 0 and col_idx == 4:
                        cell.font = Font(name="微软雅黑", size=9, color="C00000", bold=True)

                current_row += 1

            current_row += 1  # 维度间空行

        # ---- 列宽 ----
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12

    # ================================================================
    # 便捷：一步生成报告
    # ================================================================

    @staticmethod
    def save_report(results: list[dict], output_path: str, product_name: str = ""):
        """便捷方法：一步生成报告"""
        gen = ReportGenerator(product_name)
        return gen.generate(results, output_path)
