"""
统计推断驱动的费率表解析器。

不依赖关键词匹配（"终身"、"至"、"男性"），而是通过数据列的统计特征
（单调性、唯一值数、值类型、连续整数序列）自动推断每列的角色和维度映射。

支持任意保司/产品的费率表格式，无需预定义布局模板。
"""

import re
from typing import Optional, Any
from dataclasses import dataclass, field
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================
# 数据结构（保留以兼容 _build_result）
# ============================================================

@dataclass
class Region:
    """数据区段"""
    start_row: int
    end_row: int
    title_text: str = ""


@dataclass
class Layout:
    """检测到的布局结构（逐步弃用，由 col_dims dict 替代）"""
    layout_type: str = ""
    age_col: int = 1
    rate_cols: list = field(default_factory=list)
    data_start_row: int = 0
    col_dims: dict = field(default_factory=dict)
    pay_period_map: dict = field(default_factory=dict)
    gender: str = ""
    gender_map: dict = field(default_factory=dict)
    period_map: dict = field(default_factory=dict)
    plan_label: str = ""
    period: str = ""


# ============================================================
# 工具函数（保留 _clean_int, _clean_float）
# ============================================================

AGE_ZERO_PAT = re.compile(r'出生|不满1周|未满1周|0岁|0周|新生儿')


def _clean_int(val: Any) -> Optional[int]:
    if val is None: return None
    try:
        if isinstance(val, str):
            s = val.strip()
            if not s: return None
            if AGE_ZERO_PAT.search(s): return 0
            val = s.rstrip("年")
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _clean_float(val: Any) -> Optional[float]:
    if val is None: return None
    try:
        if isinstance(val, str):
            val = val.replace(",", "").replace("，", "").strip()
        return float(val)
    except (ValueError, TypeError): return None


# ============================================================
# 解析器 — 统计推断引擎
# ============================================================

class RateTableParser:
    """统计推断驱动的费率表解析器。不依赖关键词匹配。"""

    def parse(self, file_path: str) -> dict:
        wb = load_workbook(file_path, data_only=True)
        sheets = wb.sheetnames

        all_rows = []
        product_name = ""
        fee_unit = 1000
        fee_rule = "四舍五入保留2位小数"
        data_type_code = "1"
        meta = {"product_name": "", "data_type": "1", "fee_unit": 1000,
                "fee_rule": "四舍五入保留2位小数"}

        # 先检查是否有「产品信息」Sheet
        meta = self._read_metadata(wb)

        # 非费率表的 Sheet 名关键词（系数表、现金价值表等不解析）
        _NON_RATE_KEYWORDS = ("系数", "现金价值", "给付表", "示例", "演示", "说明")

        for sn in sheets:
            ws = wb[sn]
            if sn in ("产品信息", "产品说明", "info", "说明"):
                continue
            if any(kw in sn for kw in _NON_RATE_KEYWORDS):
                continue

            # === Step 1: 找年龄列 + 第一个数据行 ===
            age_col, data_start = self._find_first_age_sequence(ws)
            if age_col is None:
                continue

            # === Step 2: 检测区段（age 序列断点） ===
            sections = self._detect_sections(ws, age_col)
            if not sections:
                continue

            # === Step 3: 提取元数据 ===
            sheet_meta = self._extract_metadata(ws, age_col, data_start, file_path)

            # 标题含非费率关键词则跳过（如"系数表""现金价值"等）
            sheet_title = sheet_meta.get("product_name", "") + sn
            if any(kw in sheet_title for kw in _NON_RATE_KEYWORDS):
                continue

            if not product_name and sheet_meta.get("product_name"):
                product_name = sheet_meta["product_name"]
            if sheet_meta.get("fee_unit") and sheet_meta["fee_unit"] != 1000:
                fee_unit = sheet_meta["fee_unit"]
            if sheet_meta.get("fee_rule"):
                fee_rule = sheet_meta["fee_rule"]
            # 表头文本猜测仅在产品信息 Sheet 未指定时生效
            if sheet_meta.get("data_type_code") and sheet_meta["data_type_code"] != "1":
                if meta["data_type"] == "1":
                    data_type_code = sheet_meta["data_type_code"]

            # === Step 4: 处理每个区段 ===
            title_gender = sheet_meta.get("title_gender", "")
            title_period = sheet_meta.get("title_period", "")
            for sec_start, sec_end, sec_label in sections:
                rate_cols = self._find_rate_columns(ws, age_col, sec_start)
                if not rate_cols:
                    continue

                col_dims = self._infer_header_hierarchy(
                    ws, age_col, rate_cols, sec_start)

                # 如果表头无性别/期间，从标题元数据注入
                self._apply_title_dims(col_dims, title_gender, title_period)

                rows = self._extract_section_data(
                    ws, age_col, rate_cols, col_dims,
                    sec_start, sec_end, sec_label)
                all_rows.extend(rows)

        wb.close()

        # 合并元数据
        if product_name:
            meta["product_name"] = product_name
        if fee_unit != 1000:
            meta["fee_unit"] = fee_unit
        meta["fee_rule"] = fee_rule
        if data_type_code != "1":
            meta["data_type"] = data_type_code

        return self._build_result(all_rows, product_name, meta,
                                  [s for s in sheets
                                   if s not in ("产品信息", "产品说明", "info", "说明")],
                                  file_path)

    # ================================================================
    # Step 1: 找年龄列 + 第一个数据行
    # ================================================================

    def _parse_age(self, v):
        """解析年龄值：整数或年龄0的文本表述。非整数float返回None。"""
        if v is None: return None
        if isinstance(v, (int, float)):
            if isinstance(v, float) and v != int(v): return None
            return int(v)
        s = str(v).strip()
        if not s: return None
        if AGE_ZERO_PAT.search(s): return 0
        try: return int(float(s))
        except: return None

    def _find_first_age_sequence(self, ws: Worksheet):
        """
        扫描每个候选列，找到最早出现的有效年龄连续序列（≥5个）。
        先尝试 ≤1 岁起始（儿童产品），找不到则接受任意起始年龄（成人产品）。
        返回 (age_col, data_start_row)，无则 (None, None)。
        """
        max_row = ws.max_row
        max_col = ws.max_column

        def _find_best(max_start_age: int):
            best_col, best_start = None, float('inf')
            for col in range(1, max_col + 1):
                for row in range(1, max_row + 1):
                    a = self._parse_age(ws.cell(row=row, column=col).value)
                    if a is not None and a <= max_start_age:
                        expected = a
                        seq_len = 0
                        for r in range(row, max_row + 1):
                            v = self._parse_age(ws.cell(row=r, column=col).value)
                            if v == expected:
                                seq_len += 1; expected += 1
                            elif v is None:
                                continue
                            else:
                                break
                        if seq_len >= 5 and row < best_start:
                            best_start = row
                            best_col = col
                        break  # 每列只看第一个序列
            return best_col, best_start

        # 第一遍：≤1 岁起始（儿童产品）
        best_col, best_start = _find_best(1)
        if best_col is not None:
            return best_col, int(best_start)

        # 第二遍：放宽到任意起始年龄（成人产品，年龄从 18+ 起）
        best_col, best_start = _find_best(999)
        if best_col is None:
            return None, None
        return best_col, int(best_start)

    # ================================================================
    # Step 2: 区段检测
    # ================================================================

    def _detect_sections(self, ws: Worksheet, age_col: int):
        """
        基于 age 列数据断点检测区段。
        支持儿童产品（≤1岁起始）和成人产品（任意起始年龄）。
        返回 [(start_row, end_row, label)]。
        """
        max_row = ws.max_row

        def _collect_sequences(start_age_limit: int):
            sequences = []
            row = 1
            while row <= max_row:
                a = self._parse_age(ws.cell(row=row, column=age_col).value)
                if a is not None and a <= start_age_limit:
                    seq_start = row
                    expected = a
                    seq_end = row
                    for r in range(row, max_row + 1):
                        v = self._parse_age(ws.cell(row=r, column=age_col).value)
                        if v == expected:
                            seq_end = r; expected += 1
                        elif v is None:
                            continue
                        else:
                            break
                    if expected - a >= 5:  # 至少5个年龄
                        sequences.append((seq_start, seq_end))
                    row = seq_end + 1
                else:
                    row += 1
            return sequences

        # 先尝试 ≤1 起始，找不到则放宽到任意年龄
        sequences = _collect_sequences(1)
        if not sequences:
            sequences = _collect_sequences(999)

        # 为每个序列找标签
        sections = []
        for i, (seq_start, seq_end) in enumerate(sequences):
            label = self._extract_section_label(ws, age_col, seq_start)
            sections.append((seq_start, seq_end, label))
        return sections

    def _extract_section_label(self, ws: Worksheet, age_col: int,
                                seq_start: int) -> str:
        """从 age 序列开始行之前的行中提取区段描述文本。"""
        for r in range(seq_start - 1, max(0, seq_start - 21), -1):
            # 检查 age 列和 col 1
            for col in (age_col, 1):
                v = str(ws.cell(row=r, column=col).value or "").strip()
                if not v or len(v) < 8:
                    continue
                # 排除单位行、表头行、注释行
                if re.match(r'^[（(]?每\d+', v): continue
                if v.startswith('注：') or v.startswith('注:'): continue
                if '单位' in v and len(v) < 10: continue
                if '交费' in v and '年龄' in v: continue
                return v
        return ""

    # ================================================================
    # Step 3: 元数据提取
    # ================================================================

    def _read_metadata(self, wb) -> dict:
        """从'产品信息' Sheet 读取（保留兼容）。"""
        meta = {"product_name": "", "data_type": "1", "fee_unit": 1000,
                "fee_rule": "四舍五入保留2位小数"}
        for name in ["产品信息", "产品说明"]:
            if name in wb.sheetnames:
                ws = wb[name]
                for row in range(1, ws.max_row + 1):
                    key = str(ws.cell(row=row, column=1).value or "").strip()
                    val = str(ws.cell(row=row, column=2).value or "").strip()
                    if key == "product_name": meta["product_name"] = val
                    elif key == "data_type": meta["data_type"] = val.split("_")[0] if "_" in val else val
                    elif key == "fee_unit":
                        try: meta["fee_unit"] = int(float(val))
                        except: pass
                    elif key == "fee_rule": meta["fee_rule"] = val.split("_", 1)[1] if "_" in val else val
                break
        return meta

    def _extract_metadata(self, ws: Worksheet, age_col: int,
                           data_start: int, file_path: str) -> dict:
        """从表头区域提取产品名、费率单位、数据方向、性别、保险期间。"""
        texts = []
        # 收集 data_start 之前的描述性文本
        for row in range(1, data_start):
            for col in (1, age_col):
                v = str(ws.cell(row=row, column=col).value or "").strip()
                if v and len(v) > 5:
                    texts.append((row, col, v))

        result = {"product_name": "", "fee_unit": 0, "fee_rule": "",
                  "data_type_code": "1", "title_gender": "", "title_period": ""}

        # 产品名：最长的非维度文本
        candidates = []
        for _, _, t in texts:
            if len(t) > 10 and not re.match(r'^[（(]?每\d+', t):
                if '交费' not in t and '年龄' not in t and '若投保' not in t:
                    if '注：' not in t:
                        candidates.append(t)
        if candidates:
            result["product_name"] = max(candidates, key=len)

        # 从所有文本中提取性别、期间、单位、方向
        all_text = " ".join(t for _, _, t in texts)
        # 也检查 data_start 行（标题可能跨到数据首行）
        for col in (1, age_col):
            v = str(ws.cell(row=data_start, column=col).value or "").strip()
            if v:
                all_text += " " + v

        # 性别：标题中的"（男性"/"（女性"
        gm = re.search(r'[（(]\s*(男性|女性)', all_text)
        if gm:
            result["title_gender"] = "男" if gm.group(1) == "男性" else "女"

        # 保险期间：标题中的"保险期间终身"/"保险期间：30年"等
        pm = re.search(r'保险期间[：:\s]*(终身|至\d+周?岁|[0-9]+年)', all_text)
        if pm:
            result["title_period"] = pm.group(0).replace("保险期间", "").replace("：", "").replace(":", "").strip()

        # 费率单位
        m = re.search(r'每(\d[\d,]*)元', all_text)
        if m:
            result["fee_unit"] = int(m.group(1).replace(",", ""))

        # 加费规则
        m = re.search(r'(四舍五入|保留\d+位小数)', all_text)
        if m:
            result["fee_rule"] = m.group(0)

        # 数据方向：「每X元…保险费」→保费算保额，「每X元基本保险金额」→保额算保费
        if re.search(r'每\d+元.*保险费', all_text):
            result["data_type_code"] = "2"

        # 如果没找到产品名，尝试从文件名提取
        if not result["product_name"]:
            import os
            basename = os.path.splitext(os.path.basename(file_path))[0]
            basename = re.sub(r'_\d{8}_\d{6}$', '', basename)
            basename = basename.replace('_', '')
            if len(basename) > 3:
                result["product_name"] = basename

        return result

    # ================================================================
    # Step 4a: 费率列识别
    # ================================================================

    def _find_rate_columns(self, ws: Worksheet, age_col: int,
                            data_start: int) -> list:
        """从第一个数据行识别费率列：正浮点数且非年龄列。"""
        max_col = ws.max_column
        rate_cols = []
        for col in range(1, max_col + 1):
            if col == age_col: continue
            v = ws.cell(row=data_start, column=col).value
            if v is None: continue
            f = _clean_float(v)
            if f is not None and f > 0:
                rate_cols.append(col)
        return rate_cols

    # ================================================================
    # Step 4b: 表头层次推断
    # ================================================================

    def _infer_header_hierarchy(self, ws: Worksheet, age_col: int,
                                 rate_cols: list, data_start: int) -> dict:
        """
        分析表头行（data_start 之前的行），通过列值跨度自动推断维度层次。
        返回 {col: {"pay_period": int, "gender": str, "period": str}}。
        """
        max_col = max(rate_cols)

        # 收集表头行（仅 rate_cols 中有值的行）
        header_rows = []
        for row in range(1, data_start):
            row_vals = {}
            for col in rate_cols:
                v = ws.cell(row=row, column=col).value
                if v is not None:
                    row_vals[col] = str(v).strip()
            if row_vals:
                header_rows.append((row, row_vals))

        # 利用 openpyxl 的合并单元格信息计算真实列跨度
        merged_spans = {}
        for mr in ws.merged_cells.ranges:
            for r in range(mr.min_row, mr.max_row + 1):
                merged_spans.setdefault(r, {})[mr.min_col] = mr.max_col

        # 对每个表头行分类并计算跨度
        classified = []
        for row_num, row_vals in header_rows:
            rtype, avg_span, value_spans = self._classify_header_row(
                row_vals, merged_spans.get(row_num, {}))
            if rtype == 'data':
                continue  # 跳过数据行（可能因合并单元格混入）
            classified.append({
                'row': row_num, 'type': rtype,
                'avg_span': avg_span, 'spans': value_spans,
            })

        # 按平均跨度降序排列 → 层次深度
        classified.sort(key=lambda h: h['avg_span'], reverse=True)

        # 构建列→维度映射
        col_dims = {col: {"pay_period": None, "gender": "", "period": ""}
                    for col in rate_cols}
        for hi in classified:
            dim_type = hi['type']
            # spans 现在是 {val: [(start, end), ...]} 支持同一值多组出现
            for val, span_list in hi['spans'].items():
                for start_col, end_col in span_list:
                    for col in range(start_col, end_col + 1):
                        if col not in col_dims:
                            continue
                        if dim_type == 'pay_period':
                            pp = self._parse_pay_period_str(val)
                            if pp is not None:
                                col_dims[col]['pay_period'] = pp
                        elif dim_type == 'gender':
                            if val in ('男', '女', '男性', '女性', 'M', 'F', 'm', 'f'):
                                col_dims[col]['gender'] = '男' if val in ('男', '男性', 'M', 'm') else '女'
                        elif dim_type == 'period':
                            col_dims[col]['period'] = val

        # 扩展维度值到未标注的子列
        self._expand_dim_values(col_dims, rate_cols)

        return col_dims

    def _classify_header_row(self, row_vals: dict, merged: dict):
        """分类一个表头行的类型并计算值跨度。支持同一值在多组列出现。"""
        sorted_cols = sorted(row_vals.keys())
        # 构建 span 列表: [(val, start, end), ...] — 允许同一值多次出现
        span_list = []
        if sorted_cols:
            current_val = row_vals[sorted_cols[0]]
            span_start = sorted_cols[0]
            for i, col in enumerate(sorted_cols):
                if row_vals[col] != current_val:
                    span_list.append((current_val, span_start, sorted_cols[i-1]))
                    current_val = row_vals[col]
                    span_start = col
            span_list.append((current_val, span_start, sorted_cols[-1]))

        # 用合并单元格信息修正跨度
        for i, (val, start, end) in enumerate(span_list):
            for merge_start, merge_end in merged.items():
                if merge_start <= start <= merge_end or merge_start == start:
                    span_list[i] = (val, start, max(end, merge_end))

        # 分组：同一值可能出现在多个不连续的列组
        from collections import defaultdict
        spans_by_val = defaultdict(list)
        for val, start, end in span_list:
            spans_by_val[val].append((start, end))

        # 计算平均跨度（跨所有列组）
        all_spans = [(s, e) for groups in spans_by_val.values() for s, e in groups]
        avg_span = sum(e - s + 1 for s, e in all_spans) / len(all_spans) if all_spans else 0

        unique_vals = list(spans_by_val.keys())
        n_unique = len(unique_vals)

        # 分类特征
        all_short = all(len(v) <= 2 for v in unique_vals)
        # 仅数字+年的组合才是交费期间，避免"年满60周岁"误判
        has_year_label = any(re.search(r'\d+\s*年', v) for v in unique_vals)
        has_decimal = any(re.search(r'\.\d', v) for v in unique_vals)
        has_long = any(len(v) > 10 for v in unique_vals)

        # 数据行特征：大量十进制数值
        if has_decimal and avg_span <= 1.1:
            return 'data', avg_span, spans_by_val

        if n_unique <= 3 and all_short and not has_year_label:
            return 'gender', avg_span, spans_by_val

        if has_year_label or (n_unique <= 15 and not has_long
                              and self._parse_pay_period_str(unique_vals[0]) is not None):
            return 'pay_period', avg_span, spans_by_val

        if has_long:
            return 'period', avg_span, spans_by_val

        return 'unknown', avg_span, spans_by_val

    def _parse_pay_period_str(self, s: str) -> Optional[int]:
        """从字符串提取交费期间（纯统计，无硬编码映射表）。"""
        if not s: return None
        # "10年交" / "10年" / "十年" / "一次交清" / "趸交"
        m = re.match(r'^(\d+)\s*年', s)
        if m: return int(m.group(1))
        # 中文数字
        cn_num = {'一':1,'二':2,'三':3,'四':4,'五':5,'六':6,'七':7,'八':8,'九':9,'十':10,
                  '十五':15,'二十':20,'二十五':25,'三十':30}
        for cn, n in cn_num.items():
            if cn in s: return n
        # "一次交清" / "趸交" / "一次性交清"
        if any(w in s for w in ('一次', '趸交', '一次性')): return 1
        try: return int(s)
        except: return None

    def _expand_dim_values(self, col_dims: dict, rate_cols: list):
        """将维度值从首列扩展到所有子列（处理合并单元格）。"""
        sorted_cols = sorted(rate_cols)
        for dim_key in ('gender', 'period', 'pay_period'):
            last_val = None
            for col in sorted_cols:
                if col not in col_dims:
                    continue
                val = col_dims[col].get(dim_key)
                if val and (dim_key != 'pay_period' or val is not None):
                    last_val = val
                elif last_val is not None:
                    col_dims[col][dim_key] = last_val

    def _apply_title_dims(self, col_dims: dict, title_gender: str,
                           title_period: str):
        """当表头无性别/期间维度时，从标题提取的值注入到所有列。"""
        if not title_gender and not title_period:
            return
        has_gender = any(d.get('gender') for d in col_dims.values())
        has_period = any(d.get('period') for d in col_dims.values())
        for col in col_dims:
            if title_gender and not has_gender:
                col_dims[col]['gender'] = title_gender
            if title_period and not has_period:
                col_dims[col]['period'] = title_period

    # ================================================================
    # Step 4c: 数据提取
    # ================================================================

    def _extract_section_data(self, ws: Worksheet, age_col: int,
                               rate_cols: list, col_dims: dict,
                               sec_start: int, sec_end: int,
                               sec_label: str) -> list[dict]:
        """从区段中提取边界值数据。"""
        # 从区段标签中提取性别/期间/承保方案（处理"（男性 保险期间终身 ...）"这种标题）
        label_gender = ""
        label_period = ""
        label_ensure_plan = "1"  # 默认标准体
        if sec_label:
            gm = re.search(r'[（(]\s*(男性|女性)', sec_label)
            if gm:
                label_gender = "男" if gm.group(1) == "男性" else "女"
            pm = re.search(r'保险期间[：:\s]*(终身|至\d+周?岁|[0-9]+年)', sec_label)
            if pm:
                label_period = pm.group(0).replace("保险期间", "").replace("：", "").replace(":", "").strip()
            # 检测优选体
            if '优选体' in sec_label:
                label_ensure_plan = "2"

        # 区段标签中的性别/期间优先于 sheet 级标题（处理多区段不同性别）
        for col in col_dims:
            if label_gender:
                col_dims[col]['gender'] = label_gender
            if label_period:
                col_dims[col]['period'] = label_period

        # 构建维度键→列的映射
        boundaries = {}  # {(pay_period, gender, period): {"col": col, ...}}
        for col in rate_cols:
            dims = col_dims.get(col, {})
            pp = dims.get('pay_period')
            gender = dims.get('gender', '')
            period = dims.get('period', '')
            if pp is None:
                continue
            key = (pp, gender, period)
            # 同一 key 的多个列，取第一个（或合并处理）
            if key not in boundaries:
                boundaries[key] = {"col": col, "min_age": None, "min_rate": None,
                                   "max_age": None, "max_rate": None}

        # 逐行读取
        skipped_rows = []  # 记录被跳过的行号（调试用）
        for row in range(sec_start, sec_end + 1):
            age = self._parse_age(ws.cell(row=row, column=age_col).value)
            if age is None:
                # 允许个别空行（如区段内分隔行、合并单元格延续行）
                all_empty = True
                for b in boundaries.values():
                    if ws.cell(row=row, column=b["col"]).value is not None:
                        all_empty = False; break
                if all_empty:
                    continue
                else:
                    # 有费率但无年龄 → 跳过该行（不截断区段）
                    skipped_rows.append(row)
                    continue

            for (pp, gender, period), b in boundaries.items():
                rate = _clean_float(ws.cell(row=row, column=b["col"]).value)
                if rate is None:
                    continue
                if b["min_age"] is None:
                    b["min_age"] = age
                    b["min_rate"] = rate
                b["max_age"] = age
                b["max_rate"] = rate

        # 转为标准行
        rows = []
        for (pp, gender, period), b in boundaries.items():
            if b["min_age"] is None:
                continue
            rows.append({
                "保障方案": sec_label,
                "ensurePlan": label_ensure_plan,
                "责任计划": 0,
                "保险期间": period,
                "交费期间": pp,
                "性别": gender,
                "最小年龄": b["min_age"],
                "最小年龄费率": b["min_rate"],
                "最大年龄": b["max_age"],
                "最大年龄费率": b["max_rate"],
            })

        if skipped_rows:
            print(f"  ⚠ 区段「{sec_label}」跳过了 {len(skipped_rows)} 行（有费率但无年龄）: 行号 {skipped_rows}")

        return rows

    # ================================================================
    # 产品结构提取
    # ================================================================

    @staticmethod
    def _extract_product_structure(all_rows: list[dict]) -> dict:
        """
        从费率表行数据中提取产品结构信息，供前端动态渲染配置项。

        Returns:
            {
                "has_optional_duties": bool,    # 是否有可选责任（方案名含"+"）
                "duty_names": [str],            # 去重的可选责任名称
                "has_preferred": bool,          # 是否有优选体方案
                "plan_count": int,              # 方案种类数
                "plan_labels": {label: index},  # 方案名→自动编号
            }
        """
        plan_labels_set = set()
        has_plus = False
        has_preferred = False
        duty_names_set = set()

        for r in all_rows:
            label = str(r.get("保障方案", "")).strip()
            if not label:
                continue
            plan_labels_set.add(label)

            if "+" in label:
                has_plus = True
            if "优选体" in label:
                has_preferred = True

            # 提取 "+XXX责任" / "+XXX保险金" 中的责任名称
            parts = label.split("+")
            for part in parts[1:]:  # 跳过第一部分（基本保险责任）
                part = part.strip()
                # 去掉尾部 "责任" / "保险金责任" 等后缀，保留核心名称
                duty_name = re.sub(r'责任$', '', part)
                if duty_name and duty_name != "全部可选" and "全部可选" not in duty_name:
                    duty_names_set.add(duty_name)

        # 按首次出现顺序排序（而非字母序）
        duty_names_ordered = []
        seen = set()
        for r in all_rows:
            label = str(r.get("保障方案", "")).strip()
            parts = label.split("+")
            for part in parts[1:]:
                part = part.strip()
                duty_name = re.sub(r'责任$', '', part)
                if duty_name and duty_name != "全部可选" and "全部可选" not in duty_name and duty_name not in seen:
                    duty_names_ordered.append(duty_name)
                    seen.add(duty_name)

        # 方案名 → 编号（按首次出现顺序）
        plan_labels_ordered = []
        for r in all_rows:
            label = str(r.get("保障方案", "")).strip()
            if label and label not in plan_labels_ordered:
                plan_labels_ordered.append(label)

        plan_labels_map = {label: idx for idx, label in enumerate(plan_labels_ordered)}

        return {
            "has_optional_duties": has_plus,
            "duty_names": duty_names_ordered,
            "has_preferred": has_preferred,
            "plan_count": len(plan_labels_set),
            "plan_labels": plan_labels_map,
        }

    # ================================================================
    # 结果组装
    # ================================================================

    def _build_result(self, all_rows: list[dict], product_name: str,
                      meta: dict, data_sheets: list[str],
                      file_path: str) -> dict:
        """组装标准返回 dict。"""
        # 维度汇总
        dims = {}
        for r in all_rows:
            for k in ["保障方案", "交费期间", "性别", "责任计划"]:
                dims.setdefault(k, set()).add(r.get(k, ""))

        # 产品结构
        product_structure = self._extract_product_structure(all_rows)

        # 给每行分配正确的责任计划编号
        plan_labels_map = product_structure["plan_labels"]
        for r in all_rows:
            label = str(r.get("保障方案", "")).strip()
            if label in plan_labels_map:
                r["责任计划"] = plan_labels_map[label]

        # 边界值列表（前端用）
        boundaries = []
        for r in all_rows:
            boundaries.append({
                "label": r["保障方案"],
                "pay": r["交费期间"],
                "gender": r["性别"],
                "period": r.get("保险期间", ""),
                "min_age": r["最小年龄"],
                "min_rate": r["最小年龄费率"],
                "max_age": r["最大年龄"],
                "max_rate": r["最大年龄费率"],
            })

        case_count = sum(2 if b["min_age"] != b["max_age"] else 1
                         for b in boundaries)

        data_type = meta.get("data_type", "1")
        data_type_label = "保费算保额" if data_type == "2" else "保额算保费"

        return {
            "format": "auto",
            "product_name": product_name or meta.get("product_name", "") or "(未标注)",
            "data_type": data_type_label,
            "data_type_code": data_type,
            "fee_unit": meta.get("fee_unit", 1000),
            "fee_rule": meta.get("fee_rule", "四舍五入保留2位小数"),
            "dims": {k: sorted(v, key=str) for k, v in dims.items()},
            "boundaries": boundaries,
            "case_count": case_count,
            "boundary_count": len(boundaries),
            "product_structure": product_structure,
            "_rows": all_rows,
            "rate_sheet": data_sheets[0] if data_sheets else "",
            "file_path": file_path,
        }


# ============================================================
# 工具函数：写边界值 Excel（供 BatchTester 读取）
# ============================================================

def write_boundary_xlsx(rows: list[dict], output_path: str, product_name: str = ""):
    """将标准行字典列表写入边界值汇总 Excel（TestCaseLoader 兼容格式）。"""
    from datetime import datetime as dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADERS = [
        "保障方案", "ensurePlan", "责任计划", "保险期间", "交费期间（年）",
        "性别", "最小年龄（岁）", "最小年龄费率", "最大年龄（岁）", "最大年龄费率",
    ]
    COL_WIDTHS = [10, 12, 10, 10, 14, 8, 14, 14, 14, 14]
    ncols = len(HEADERS)

    wb = Workbook()
    ws = wb.active
    ws.title = "边界费率汇总"

    title = f"{product_name or '保险产品'} — 边界费率汇总表  生成时间：{dt.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=1, value="保险期间：终身 | 金额单位：元（每1,000元保额）")
    c.font = Font(name="微软雅黑", size=9, color="666666")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 22

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for i, row_data in enumerate(rows):
        row_num = 4 + i
        is_even = (i % 2 == 0)
        bg = "DEEAF1" if is_even else "FFFFFF"
        row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

        values = [
            row_data.get("保障方案", ""),
            row_data.get("ensurePlan", "1"),
            row_data.get("责任计划", 0),
            row_data.get("保险期间", ""),
            row_data.get("交费期间", ""),
            row_data.get("性别", ""),
            row_data.get("最小年龄", ""),
            row_data.get("最小年龄费率", ""),
            row_data.get("最大年龄", ""),
            row_data.get("最大年龄费率", ""),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="微软雅黑", size=10, color="1F4E79",
                            bold=(col_idx <= 2))
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(rows)}"

    wb.save(output_path)
    print(f"✅ 边界值汇总已写入: {output_path}")
