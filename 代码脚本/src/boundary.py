"""
边界值汇总生成器 — 纯格式驱动，与产品解耦。
"""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import FormatProfile, ProductProfile
from .rate_reader import RateTableReader


class BoundarySummaryGenerator:
    """
    产出标准 10 列边界值汇总 Excel。

    用法:
        fmt = FormatProfile.from_yaml("formats/grid.yaml")
        gen = BoundarySummaryGenerator(fmt)
        gen.generate("费率表.xlsx", "边界值汇总.xlsx")
    """

    COLOR_STANDARD_EVEN = "DEEAF1"
    COLOR_STANDARD_ODD = "FFFFFF"
    COLOR_PREFERRED_EVEN = "E2EFDA"
    COLOR_PREFERRED_ODD = "F9FFF7"
    COLOR_STANDARD_TEXT = "1F4E79"
    COLOR_PREFERRED_TEXT = "375623"

    HEADERS = [
        "保障方案", "ensurePlan", "责任计划", "保险期间", "交费期间（年）",
        "性别", "最小年龄（岁）", "最小年龄费率", "最大年龄（岁）", "最大年龄费率",
    ]
    COL_WIDTHS = [10, 12, 10, 10, 14, 8, 14, 14, 14, 14]

    def __init__(self, format_profile: FormatProfile, product_profile: ProductProfile = None):
        self.format = format_profile
        self.product = product_profile

    @property
    def _product_name(self) -> str:
        if self.product:
            return self.product.product_name
        return self.format.format_name or "保险产品"

    @property
    def _has_preferred(self) -> bool:
        if self.product and self.product.age_limits.preferred:
            return True
        return False

    @property
    def _preferred_limits(self) -> dict:
        if self.product and self.product.age_limits.preferred:
            return self.product.age_limits.preferred
        return {}

    # ================================================================
    # 主入口
    # ================================================================

    def generate(self, rate_file: str, output_file: str, use_grouped: bool = False) -> str:
        print(f"📂 读取费率表: {rate_file}")
        reader = RateTableReader(self.format, self.product)

        if use_grouped:
            rows = reader.read_grouped(rate_file)
        else:
            rows = reader.read_all_sections(rate_file)

        if not rows:
            print("⚠ 未提取到任何数据")
            return output_file

        rows.sort(key=lambda r: (
            int(r.get("ensurePlan", 0)),
            int(r.get("责任计划", 0)),
            int(r.get("交费期间", 0)),
            r.get("性别", ""),
        ))

        print(f"📝 生成边界值汇总: {output_file}")
        wb = Workbook()

        self._write_summary_sheet(wb, rows)
        self._write_age_range_sheet(wb, rows)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(output_file)
        print(f"✅ 边界值汇总已生成: {output_file}")
        self._print_stats(rows)
        return output_file

    # ================================================================
    # Sheet 1: 边界费率汇总
    # ================================================================

    def _write_summary_sheet(self, wb: Workbook, rows: list[dict]):
        ws = wb.active
        ws.title = "边界费率汇总"

        ncols = len(self.HEADERS)

        # 标题行
        title = f"{self._product_name} — 边界费率汇总表  生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 说明行
        sections_desc = " + ".join(f"{s.label}(ensurePlan={s.ensure_plan})" for s in self.format.sections)
        note = f"保险期间：终身 | {sections_desc} | 金额单位：元（每1,000元保额）"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        c = ws.cell(row=2, column=1, value=note)
        c.font = Font(name="微软雅黑", size=9, color="666666")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 22

        # 表头行
        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        for col_idx, header in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 22

        # 数据行
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        for i, row_data in enumerate(rows):
            row_num = 4 + i
            is_preferred = row_data.get("ensurePlan") == "2"
            is_even = (i % 2 == 0)

            bg = self.COLOR_PREFERRED_EVEN if is_preferred and is_even else \
                 self.COLOR_PREFERRED_ODD if is_preferred else \
                 self.COLOR_STANDARD_EVEN if is_even else self.COLOR_STANDARD_ODD
            text_color = self.COLOR_PREFERRED_TEXT if is_preferred else self.COLOR_STANDARD_TEXT

            row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

            values = [
                row_data.get("保障方案", ""),
                row_data.get("ensurePlan", ""),
                row_data.get("责任计划", ""),
                row_data.get("保险期间", ""),
                row_data.get("交费期间", ""),
                row_data.get("性别", ""),
                row_data.get("最小年龄", ""),
                row_data.get("最小年龄费率", ""),
                row_data.get("最大年龄", ""),
                row_data.get("最大年龄费率", ""),
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = Font(name="微软雅黑", size=10, color=text_color,
                                 bold=(col_idx <= 2))
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # 列宽
        for col_idx, width in enumerate(self.COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}{3 + len(rows)}"

    # ================================================================
    # Sheet 2: 承保年龄范围说明
    # ================================================================

    def _write_age_range_sheet(self, wb: Workbook, rows: list[dict]):
        ws = wb.create_sheet("承保年龄范围说明")

        ws.merge_cells("A1:D1")
        c = ws.cell(row=1, column=1, value="各交费期间对应承保年龄范围")
        c.font = Font(name="微软雅黑", size=13, bold=True, color="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hdr_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        for col_idx, h in enumerate(["交费期间（年）", "最小承保年龄（岁）", "最大承保年龄（岁）", "备注"], 1):
            cell = ws.cell(row=3, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 从数据聚合年龄范围
        pay_range = {}
        for r in rows:
            pay = r.get("交费期间")
            if pay is None:
                continue
            if pay not in pay_range:
                pay_range[pay] = {"min": r.get("最小年龄"), "max": r.get("最大年龄")}
            else:
                ma = r.get("最小年龄")
                if ma is not None:
                    cur = pay_range[pay]["min"]
                    pay_range[pay]["min"] = ma if cur is None else min(cur, ma)
                mx = r.get("最大年龄")
                if mx is not None:
                    cur = pay_range[pay]["max"]
                    pay_range[pay]["max"] = mx if cur is None else max(cur, mx)

        for i, pay in enumerate(sorted(pay_range.keys())):
            row_num = 4 + i
            info = pay_range[pay]
            ws.cell(row=row_num, column=1, value=pay).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=2, value=info["min"]).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=3, value=info["max"]).font = Font(name="微软雅黑", size=10)
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=4, value=f"承保年龄 {info['min']}~{info['max']} 岁").font = Font(name="微软雅黑", size=10)

            if i % 2 == 0:
                fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
                for c in range(1, 5):
                    ws.cell(row=row_num, column=c).fill = fill

        # 优选体脚注
        if self._has_preferred:
            pref = self._preferred_limits
            foot_row = 4 + len(pay_range) + 1
            note = (f"注：优选体（ensurePlan=2）要求被保人年龄 {pref.get('min', 18)}~{pref.get('max', 65)} 周岁，"
                    f"0~{pref.get('min', 18) - 1} 岁及 {pref.get('max', 65) + 1} 岁以上仅适用标准体")
            ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=4)
            ws.cell(row=foot_row, column=1, value=note).font = Font(name="微软雅黑", size=9, bold=True, color="C00000")

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 16
        ws.column_dimensions["D"].width = 36

    # ================================================================
    # 统计
    # ================================================================

    def _print_stats(self, rows: list[dict]):
        standard = [r for r in rows if r.get("ensurePlan") == "1"]
        preferred = [r for r in rows if r.get("ensurePlan") != "1"]
        plans = sorted(set(r.get("责任计划") for r in rows))
        pays = sorted(set(r.get("交费期间") for r in rows))

        print(f"\n{'='*50}")
        print(f"  总记录数: {len(rows)}")
        if standard:
            print(f"  标准体:   {len(standard)} 条")
        if preferred:
            print(f"  优选体:   {len(preferred)} 条")
        print(f"  责任计划: {plans}")
        print(f"  交费期间: {pays}")
        print(f"{'='*50}\n")
