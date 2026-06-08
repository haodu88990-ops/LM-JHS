"""
费率表读取器 — 读取 Excel 费率表的结构、元数据和边界值。
"""

import re
from typing import Optional, Any
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import FormatProfile, SheetSection, ProductProfile


class RateTableMetadata:
    """费率表元数据 — 从'产品信息'Sheet 提取"""

    def __init__(self):
        self.product_name: str = ""
        self.out_product_code: str = ""
        self.data_type: str = "1"            # "1"=保额算保费, "2"=保费算保额
        self.data_type_label: str = "保额算保费"
        self.fee_unit: int = 1000
        self.fee_rule: str = "四舍五入保留2位小数"
        self.main_flag: str = "1_主险"
        self.fee_table_type: str = "1_首期保费表"

    @property
    def is_premium_to_coverage(self) -> bool:
        """是否「保费算保额」"""
        return self.data_type == "2"

    def summary(self) -> str:
        lines = [
            f"  产品名称: {self.product_name}",
            f"  产品编码: {self.out_product_code}",
            f"  算费方向: {self.data_type_label}",
            f"  算费单位: 每{self.fee_unit}元",
            f"  舍入规则: {self.fee_rule}",
        ]
        return "\n".join(lines)


class RateTableReader:
    """
    费率表读取器。

    用法:
        fmt = FormatProfile.from_yaml("formats/grid.yaml")
        reader = RateTableReader(fmt)
        meta = reader.read_metadata("费率表.xlsx")
        data = reader.read_all_sections("费率表.xlsx")
    """

    def __init__(self, format_profile: FormatProfile, product_profile: ProductProfile = None):
        self.format = format_profile
        self.layout = format_profile.layout
        self.product = product_profile

    # ================================================================
    # 元数据
    # ================================================================

    def read_metadata(self, file_path: str) -> RateTableMetadata:
        """从费率表的'产品信息'Sheet 提取元数据"""
        wb = load_workbook(file_path, data_only=True)
        meta = RateTableMetadata()

        info_sheet = None
        for name in ["产品信息", "产品说明", "说明", "info"]:
            if name in wb.sheetnames:
                info_sheet = wb[name]
                break

        if info_sheet:
            for row in range(1, info_sheet.max_row + 1):
                key = str(info_sheet.cell(row=row, column=1).value or "").strip()
                val = str(info_sheet.cell(row=row, column=2).value or "").strip()

                if key == "product_name":
                    meta.product_name = val
                elif key == "out_product_code":
                    meta.out_product_code = val
                elif key == "data_type":
                    meta.data_type = val.split("_")[0] if "_" in val else val
                    meta.data_type_label = "保费算保额" if meta.data_type == "2" else "保额算保费"
                elif key == "fee_unit":
                    try:
                        meta.fee_unit = int(float(val))
                    except ValueError:
                        pass
                elif key == "fee_rule":
                    meta.fee_rule = val.split("_", 1)[1] if "_" in val else val
                elif key == "main_flag":
                    meta.main_flag = val
                elif key == "fee_table_type":
                    meta.fee_table_type = val

        wb.close()
        return meta

    # ================================================================
    # 主入口
    # ================================================================

    def read_all_sections(self, file_path: str) -> list[dict]:
        """读取格式配置中定义的所有 Sheet/区段"""
        wb = load_workbook(file_path, data_only=True)
        all_rows = []

        for section in self.format.sections:
            if section.sheet not in wb.sheetnames:
                print(f"  ⚠ 跳过不存在的 Sheet: {section.sheet}")
                continue
            ws = wb[section.sheet]
            rows = self._extract_section(ws, section)
            all_rows.extend(rows)
            print(f"  ✓ {section.sheet} ({section.label}): {len(rows)} 条")

        wb.close()
        return all_rows

    # ================================================================
    # 探索
    # ================================================================

    def explore(self, file_path: str) -> dict:
        if not file_path:
            raise ValueError("explore() 需要指定费率表文件路径")
        wb = load_workbook(file_path, data_only=True)
        result = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            info = {
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "rows": {},
            }
            for row in range(1, min(11, ws.max_row + 1)):
                cells = []
                for col in range(1, min(16, ws.max_column + 1)):
                    val = ws.cell(row=row, column=col).value
                    if val is not None:
                        cells.append((col, val))
                if cells:
                    info["rows"][row] = cells
            result[sheet_name] = info
        wb.close()
        return result

    # ================================================================
    # 内部提取
    # ================================================================

    def _extract_section(self, ws: Worksheet, section: SheetSection) -> list[dict]:
        if self.layout.layout_type == "grid":
            return self._extract_grid(ws, section)
        return self._extract_column(ws, section)

    def _extract_column(self, ws: Worksheet, section: SheetSection) -> list[dict]:
        layout = self.layout
        results = []
        start_col = section.column_start or layout.rate_columns_start
        end_col = section.column_end or layout.rate_columns_end or ws.max_column

        for col in range(start_col, end_col + 1):
            plan_val = ws.cell(row=layout.header_rows["plan"], column=col).value
            period_val = ws.cell(row=layout.header_rows["period"], column=col).value
            pay_val = ws.cell(row=layout.header_rows["pay_period"], column=col).value
            gender_val = ws.cell(row=layout.header_rows["gender"], column=col).value

            period = self._clean_str(period_val)
            gender = self._clean_str(gender_val)
            if gender not in ("男", "女") or not period:
                continue

            pay_period = self._normalize_pay_period(pay_val)
            if pay_period is None:
                continue

            if section.plan_override is not None:
                plan = section.plan_override
            elif section.plan_map:
                plan_str = self._clean_str(plan_val)
                plan = section.plan_map.get(plan_str, self._clean_int(plan_val))
                if plan is None:
                    continue
            else:
                plan = self._clean_int(plan_val)
                if plan is None:
                    continue

            data_start = layout.data_start_row
            data_end = layout.data_end_row or ws.max_row
            age_col = layout.age_column

            min_age, min_rate = None, None
            max_age, max_rate = None, None

            for row in range(data_start, data_end + 1):
                age = self._clean_int(ws.cell(row=row, column=age_col).value)
                rate = self._clean_float(ws.cell(row=row, column=col).value)
                if age is not None and rate is not None:
                    if min_age is None:
                        min_age, min_rate = age, rate
                    max_age, max_rate = age, rate

            if min_age is None:
                continue

            results.append({
                "保障方案": section.label,
                "ensurePlan": section.ensure_plan,
                "责任计划": plan,
                "保险期间": period,
                "交费期间": pay_period,
                "性别": gender,
                "最小年龄": min_age,
                "最小年龄费率": min_rate,
                "最大年龄": max_age,
                "最大年龄费率": max_rate,
            })

        return results

    def _extract_grid(self, ws: Worksheet, section: SheetSection) -> list[dict]:
        layout = self.layout
        pp_row = layout.pay_period_row
        g_row = layout.gender_row
        age_col = layout.age_column
        data_start = layout.data_start_row
        data_end = layout.data_end_row or ws.max_row
        start_col = section.column_start or layout.rate_columns_start
        end_col = section.column_end or layout.rate_columns_end or ws.max_column
        results = []

        for col in range(start_col, end_col + 1):
            pp_raw = ws.cell(row=pp_row, column=col).value
            gender_raw = ws.cell(row=g_row, column=col).value
            pay_period = self._normalize_pay_period(pp_raw)
            gender = self._clean_str(gender_raw)

            if pay_period is None:
                continue
            if gender in ("M", "F"):
                gender = "男" if gender == "M" else "女"
            if gender not in ("男", "女"):
                continue

            min_age, min_rate = None, None
            max_age, max_rate = None, None

            for row in range(data_start, data_end + 1):
                age = self._clean_int(ws.cell(row=row, column=age_col).value)
                rate = self._clean_float(ws.cell(row=row, column=col).value)
                if age is not None and rate is not None:
                    if min_age is None:
                        min_age, min_rate = age, rate
                    max_age, max_rate = age, rate

            if min_age is None:
                continue

            plan = section.plan_override if section.plan_override is not None else 0
            period = layout.period_override or "终身"

            results.append({
                "保障方案": section.label,
                "ensurePlan": section.ensure_plan,
                "责任计划": plan,
                "保险期间": period,
                "交费期间": pay_period,
                "性别": gender,
                "最小年龄": min_age,
                "最小年龄费率": min_rate,
                "最大年龄": max_age,
                "最大年龄费率": max_rate,
            })

        return results

    # ================================================================
    # 工具
    # ================================================================

    def _normalize_pay_period(self, raw_value) -> Optional[int]:
        if raw_value is None:
            return None
        if isinstance(raw_value, (int, float)):
            return int(raw_value)
        s = str(raw_value).strip()
        if not s:
            return None
        if self.product:
            result = self.product.normalize_pay_period(s)
            if result is not None:
                return result
        try:
            return int(s)
        except ValueError:
            pass
        if s in ("趸交", "一次性交清"):
            return 1
        m = re.match(r'^(\d+)年?$', s)
        if m:
            return int(m.group(1))
        return None

    @staticmethod
    def _clean_int(val: Any) -> Optional[int]:
        if val is None:
            return None
        try:
            if isinstance(val, str):
                val = val.strip().rstrip("年")
                if not val:
                    return None
            return int(float(val))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _clean_float(val: Any) -> Optional[float]:
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _clean_str(val: Any) -> Optional[str]:
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None
