#!/usr/bin/env python3
"""
保费测算工具 — 一键运行。

用法:
    python run.py <费率表.xlsx> [选项]

流程全自动:
    1. 自动识别费率表格式（grid / column）
    2. 自动读取产品信息（算费方向、费率单位...）
    3. 展示测算条件与边界值
    4. 交互式询问 serialNo
    5. API 批量测算
    6. 输出 Excel 对比报告

示例:
    python run.py 费率表.xlsx
    python run.py 费率表.xlsx --serial-no 1491455820292947968
    python run.py 费率表.xlsx --smoke
"""

import sys
import os
import argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from src.config import (
    FormatProfile, ProductProfile, RateLayoutConfig, SheetSection
)
from src.rate_reader import RateTableReader
from src.boundary import BoundarySummaryGenerator
from src.tester import BatchTester
from src.reporter import ReportGenerator


# ================================================================
# 内置默认值（费率表里有产品信息的就用产品信息，没有的用这些）
# ================================================================

DEFAULT_PRODUCT = {
    "product_name": "",
    "api": {
        "base_url": "https://kfzxtb.lmbaoxian.com:13080",
        "login_endpoint": "/broker/api/user/login.html",
        "age_rate_endpoint": "/broker/api/prospectus/saveCustomer.html",
        "plan_rate_endpoint": "/broker/api/prospectus/saveProductExt.html",
        "verify_ssl": False,
        "timeout": 30,
    },
    "credentials": {
        "account": "15856990088",
        "password_md5": "dc483e80a7a0bd9ef71d8cf973673924",
    },
    "product": {
        "product_id": "991452",
        "company_id": "100080",
    },
    "defaults": {
        "insurant_id": 85320,
        "insurant_occ_level": 1,
        "insurant_social_insurance": "1",
        "policy_holder_id": 85321,
        "policy_holder_age": 30,
        "policy_holder_sex": "1",
        "dividend_draw_type": "2",
        "request_type": "prospectus",
    },
    "mappings": {
        "gender": {"男": "1", "女": "2"},
        "ensure_period": {"终身": "TO105"},
        "pay_mode": {"single": "1", "annual": "5"},
        "pay_period": "direct",
    },
    "plans": {"type": "simple"},
    "age_limits": {},
    "test": {
        "amount": {"min": 1000000, "max": 5000000, "step": 1000},
        "tolerance": 0.01,
        "throttle": {"interval": 5, "sleep": 0.3},
    },
    "output": {},
}


# ================================================================
# 格式自动检测
# ================================================================

def detect_format(file_path: str) -> str:
    """
    自动检测费率表格式。

    策略：
    1. 检查是否有 '产品信息' + '费率' Sheet → grid
    2. 检查 '费率' Sheet 的结构 → grid 或 column
    3. 检查是否有 '标准体费率表' Sheet → column
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path, data_only=True)
    sheets = wb.sheetnames

    # 有产品信息 Sheet → 很可能是标准格式
    has_info = any(s in sheets for s in ["产品信息", "产品说明"])

    # 检查费率 Sheet
    rate_sheet = None
    for s in sheets:
        if "费率" in s and "产品" not in s:
            rate_sheet = s
            break
    if not rate_sheet and len(sheets) > 0:
        rate_sheet = sheets[-1]  # 最后一个 Sheet 通常是数据

    if rate_sheet:
        ws = wb[rate_sheet]
        # Grid 特征: Row 3 有 PREMIUM / pay_period 字样
        row3_first = str(ws.cell(row=3, column=1).value or "").strip().upper()
        row3_second = str(ws.cell(row=3, column=2).value or "").strip()
        if row3_first in ("PREMIUM", "保费") or "pay_period" in row3_second.lower():
            wb.close()
            return "grid"

    # 如果有多 Sheet 且包含"费率表"字样 → column
    for s in sheets:
        if "费率表" in s:
            wb.close()
            return "column"

    # 默认尝试 column
    wb.close()
    return "column"


def build_format(file_path: str) -> FormatProfile:
    """根据检测结果构建 FormatProfile"""
    fmt_type = detect_format(file_path)
    print(f"  识别格式: {fmt_type}")

    if fmt_type == "grid":
        return FormatProfile(
            format_name="自动识别",
            format_type="grid",
            layout=RateLayoutConfig(
                layout_type="grid",
                pay_period_row=3, gender_row=4,
                data_start_row=5, age_column=2, rate_columns_start=3,
            ),
            sections=[SheetSection(sheet="费率", label="标准体", ensure_plan="1", plan_override=0)],
        )
    else:
        return FormatProfile(
            format_name="自动识别",
            format_type="column",
            layout=RateLayoutConfig(
                layout_type="column",
                header_rows={"plan": 4, "period": 5, "pay_period": 6, "gender": 7},
                data_start_row=8, age_column=1, rate_columns_start=2,
            ),
            sections=[SheetSection(sheet="标准体费率表", label="标准体", ensure_plan="1")],
        )


# ================================================================
# 交互
# ================================================================

def prompt_serial_no() -> str:
    print()
    print("🔑 请输入模拟测算的 serialNo（方案序列号）：")
    print("   此参数从业务系统获取，没有它无法进行保费测算")
    print()
    while True:
        val = input("   serialNo: ").strip()
        if val:
            return val
        print("   ⚠ serialNo 不能为空")


def prompt_account() -> tuple:
    """如果默认账号不对，让用户改"""
    print()
    print("📧 登录账号（回车使用默认）:")
    account = input(f"   账号 [{DEFAULT_PRODUCT['credentials']['account']}]: ").strip()
    if not account:
        account = DEFAULT_PRODUCT['credentials']['account']
    password = input(f"   密码(已加密) [{DEFAULT_PRODUCT['credentials']['password_md5'][:8]}...]: ").strip()
    if not password:
        password = DEFAULT_PRODUCT['credentials']['password_md5']
    return account, password


# ================================================================
# 主流程
# ================================================================

def main():
    parser = argparse.ArgumentParser(
        description="保费测算工具 — 一键对比费率表与 API 测算结果",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("rate_file", help="费率表 Excel 路径")
    parser.add_argument("--serial-no", "-s", help="方案序列号")
    parser.add_argument("--proposal-id", default="", help="投保单号")
    parser.add_argument("--output", "-o", help="报告输出路径")
    parser.add_argument("--account", help="登录账号")
    parser.add_argument("--smoke", action="store_true", help="冒烟测试（前5条）")
    parser.add_argument("--smoke-count", type=int, default=5)
    args = parser.parse_args()

    if not os.path.exists(args.rate_file):
        print(f"❌ 文件不存在: {args.rate_file}")
        sys.exit(1)

    rate_file = os.path.abspath(args.rate_file)

    # ================================================================
    # Phase 1: 解析费率表
    # ================================================================
    print()
    print("=" * 60)
    print("  Phase 1: 解析费率表")
    print("=" * 60)
    print(f"  文件: {rate_file}")

    fmt = build_format(rate_file)

    # 先读元数据
    reader = RateTableReader(fmt)
    meta = reader.read_metadata(rate_file)

    # 用元数据补全产品配置
    product_data = dict(DEFAULT_PRODUCT)
    if meta.product_name:
        product_data["product_name"] = meta.product_name
    product = ProductProfile.from_dict(product_data)

    # 用产品映射重新创建 reader
    reader = RateTableReader(fmt, product)

    print(meta.summary())

    # 提取边界值
    rows = reader.read_all_sections(rate_file)

    # 分析条件
    from collections import defaultdict
    dims = defaultdict(set)
    for r in rows:
        dims["保障方案"].add(r.get("保障方案", ""))
        dims["交费期间"].add(r.get("交费期间"))
        dims["性别"].add(r.get("性别"))
        dims["年龄范围"].add(f"{r.get('最小年龄')}~{r.get('最大年龄')}岁")

    print(f"\n  测算条件:")
    for k, v in dims.items():
        print(f"    {k}: {sorted(v, key=str)}")

    print(f"\n  边界值 ({len(rows)} 条):")
    for r in rows:
        print(f"    {r['保障方案']:8s} | 交{r['交费期间']:2d}年 | {r['性别']} | "
              f"{r['最小年龄']:2d}岁({r['最小年龄费率']}‰) → {r['最大年龄']:2d}岁({r['最大年龄费率']}‰)")

    case_count = sum(2 if r['最小年龄'] != r['最大年龄'] else 1 for r in rows)
    print(f"\n  预计测试用例: {case_count} 条")

    # ================================================================
    # Phase 2: 参数确认
    # ================================================================
    print()
    print("=" * 60)
    print("  Phase 2: 参数确认")
    print("=" * 60)

    serial_no = args.serial_no or prompt_serial_no()
    proposal_id = args.proposal_id

    # 登录账号：命令行传了就直接用，否则交互问
    account = args.account or DEFAULT_PRODUCT["credentials"]["account"]
    password = DEFAULT_PRODUCT["credentials"]["password_md5"]
    if args.serial_no:
        # 命令行模式：不交互，直接用默认
        pass
    else:
        account, password = prompt_account()

    product.credentials.account = account
    product.credentials.password_md5 = password

    print(f"\n  ✅ serialNo: {serial_no}")
    print(f"  ✅ API:     {product.api.base_url}")
    print(f"  ✅ 产品:    {product.product_name or '(费率表未标注)'}")
    print(f"  ✅ 用例数:  {case_count}")

    if not args.serial_no:
        print()
        confirm = input("  开始测算? [Y/n]: ").strip().lower()
        if confirm and confirm != "y":
            print("  已取消")
            return

    # ================================================================
    # Phase 3: API 测算
    # ================================================================
    print()
    print("=" * 60)
    print("  Phase 3: API 保费测算")
    print("=" * 60)

    tmp_boundary = os.path.join(os.path.dirname(rate_file), "_boundary_tmp.xlsx")
    gen = BoundarySummaryGenerator(fmt, product)
    gen.generate(rate_file=rate_file, output_file=tmp_boundary)

    tester = BatchTester(product, serial_no=serial_no, proposal_id=proposal_id)
    results = tester.run(
        boundary_file=tmp_boundary,
        smoke_only=args.smoke,
        smoke_count=args.smoke_count,
    )

    if tmp_boundary and os.path.exists(tmp_boundary):
        os.remove(tmp_boundary)

    if not results:
        return

    # ================================================================
    # Phase 4: 报告
    # ================================================================
    print()
    print("=" * 60)
    print("  Phase 4: 生成报告")
    print("=" * 60)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = args.output or f"保费测算报告_{ts}.xlsx"
    reporter = ReportGenerator(product.product_name or "保险产品")
    reporter.generate(results, output)

    print(f"\n  ✅ 报告: {output}")
    print()


if __name__ == "__main__":
    main()
