#!/usr/bin/env python3
"""
探索费率表结构

快速查看 Excel 费率表的 Sheet 组成和布局，帮助编写产品配置文件。

用法:
    python explore.py <费率表.xlsx>

示例:
    python explore.py 费率表.xlsx
"""

import sys
import os

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook


def main():
    if len(sys.argv) < 2:
        print("用法: python explore.py <费率表.xlsx>")
        sys.exit(1)

    file_path = sys.argv[1]
    if not os.path.exists(file_path):
        print(f"❌ 文件不存在: {file_path}")
        sys.exit(1)

    print(f"📂 探索费率表: {file_path}\n")
    wb = load_workbook(file_path, data_only=True)

    print(f"Sheet 数量: {len(wb.sheetnames)}")
    print(f"Sheet 名称: {wb.sheetnames}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"{'='*70}")
        print(f"Sheet: {sheet_name}")
        print(f"  范围: {ws.dimensions}")
        print(f"  行数: {ws.max_row}, 列数: {ws.max_column}")

        # 显示前 10 行
        print(f"\n  前 10 行内容 (每行最多显示前 20 列):")
        for row in range(1, min(11, ws.max_row + 1)):
            cells = []
            for col in range(1, min(21, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    cells.append((col, str(val)[:50]))
            if cells:
                cell_str = " | ".join(f"[C{col}] {v}" for col, v in cells)
                print(f"    Row {row:2d}: {cell_str}")
            else:
                print(f"    Row {row:2d}: (空)")

        # 显示最后 3 行
        if ws.max_row > 10:
            print(f"\n  最后 3 行内容 (行 {ws.max_row - 2} ~ {ws.max_row}):")
            for row in range(max(11, ws.max_row - 2), ws.max_row + 1):
                cells = []
                for col in range(1, min(21, ws.max_column + 1)):
                    val = ws.cell(row=row, column=col).value
                    if val is not None:
                        cells.append((col, str(val)[:50]))
                if cells:
                    cell_str = " | ".join(f"[C{col}] {v}" for col, v in cells)
                    print(f"    Row {row:2d}: {cell_str}")

        print()

    wb.close()
    print("✅ 探索完成。根据以上信息编写产品配置文件。")


if __name__ == "__main__":
    main()
