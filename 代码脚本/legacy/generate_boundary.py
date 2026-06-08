# -*- coding: utf-8 -*-
"""
从费率表.xlsx 提取标准体+优选体边界值，生成「瑞泰鸿利致享版边界值汇总.xlsx」
新增 ensurePlan 字段：1=标准体，2=优选体
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

SRC = r'f:\workbuddy\test_Jhs\费率表.xlsx'
OUT = r'f:\workbuddy\test_Jhs\瑞泰鸿利致享版边界值汇总.xlsx'

SHEET_MAP = [
    ('标准体费率表', '1', '标准体'),
    ('优选体费率表', '2', '优选体'),
]


def extract_boundary(ws):
    """从费率 Sheet 提取所有组合的边界年龄+费率，返回 list of dict"""
    rows = []
    for c in range(2, ws.max_column + 1):
        plan   = ws.cell(row=4, column=c).value
        period = ws.cell(row=5, column=c).value
        pay    = ws.cell(row=6, column=c).value
        gender = str(ws.cell(row=7, column=c).value or '').strip()
        if plan is None or gender not in ('男', '女'):
            continue

        min_age = max_age = min_rate = max_rate = None
        for r in range(8, ws.max_row + 1):
            age = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=c).value
            if age is not None and val is not None and val != '':
                try:
                    val = float(val)
                except (TypeError, ValueError):
                    continue
                if min_age is None:
                    min_age, min_rate = int(age), val
                max_age, max_rate = int(age), val

        if min_age is None:
            continue
        rows.append({
            '责任计划': int(plan),
            '保险期间': str(period),
            '交费期间': int(pay),
            '性别': gender,
            '最小年龄': min_age,
            '最小年龄费率': min_rate,
            '最大年龄': max_age,
            '最大年龄费率': max_rate,
        })
    return rows


def write_summary_sheet(wb, all_rows):
    """写「边界费率汇总」Sheet（标准体+优选体合并，新增 ensurePlan 列）"""
    ws = wb.active
    ws.title = '边界费率汇总'

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font  = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
    title_fill  = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    std_fill    = PatternFill(fill_type='solid', fgColor='2E75B6')   # 标准体表头
    opt_fill    = PatternFill(fill_type='solid', fgColor='375623')   # 优选体表头（深绿）
    row_std_e   = PatternFill(fill_type='solid', fgColor='DEEAF1')   # 标准体偶数行
    row_std_o   = PatternFill(fill_type='solid', fgColor='FFFFFF')
    row_opt_e   = PatternFill(fill_type='solid', fgColor='E2EFDA')   # 优选体偶数行
    row_opt_o   = PatternFill(fill_type='solid', fgColor='F9FFF7')

    # 标题
    ws.merge_cells('A1:I1')
    ws['A1'] = f'瑞泰鸿利传世（致享版）终身寿险 — 标准体 & 优选体 边界费率汇总表  生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 32

    # 说明行
    ws.merge_cells('A2:I2')
    ws['A2'] = '保险期间：终身 | ensurePlan: 1=标准体, 2=优选体 | 各交费期间对应不同最大承保年龄 | 金额单位：元（每1,000元保额）'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='595959')
    ws['A2'].fill = PatternFill(fill_type='solid', fgColor='DEEAF1')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 18

    # 表头
    headers = ['保障方案', 'ensurePlan', '责任计划', '保险期间', '交费期间（年）', '性别',
               '最小年龄（岁）', '最小年龄费率', '最大年龄（岁）', '最大年龄费率']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = std_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 32

    # 数据行
    ri = 4
    for row in all_rows:
        plan_code = row['ensurePlan']
        plan_name = row['保障方案']
        is_opt = (plan_code == '2')
        cnt = ri - 4
        if is_opt:
            bg = row_opt_e if cnt % 2 == 0 else row_opt_o
        else:
            bg = row_std_e if cnt % 2 == 0 else row_std_o

        vals = [plan_name, plan_code, row['责任计划'], row['保险期间'],
                row['交费期间'], row['性别'],
                row['最小年龄'], row['最小年龄费率'],
                row['最大年龄'], row['最大年龄费率']]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = Font(name='微软雅黑', size=9,
                             bold=(ci <= 2),
                             color='375623' if is_opt else '1F4E79')
            cell.fill = bg
            cell.alignment = center
            cell.border = border
        ws.row_dimensions[ri].height = 18
        ri += 1

    # 列宽
    col_widths = [10, 12, 10, 10, 14, 8, 14, 14, 14, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    std_count = sum(1 for r in all_rows if r['ensurePlan'] == '1')
    opt_count = sum(1 for r in all_rows if r['ensurePlan'] == '2')
    print(f'  边界费率汇总 Sheet: 标准体{std_count}条 + 优选体{opt_count}条 = 共{len(all_rows)}条')


def write_age_range_sheet(wb):
    """写「承保年龄范围说明」Sheet"""
    ws = wb.create_sheet('承保年龄范围说明')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2E75B6')
    title_font  = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    title_fill  = PatternFill(fill_type='solid', fgColor='1F4E79')

    ws.merge_cells('A1:D1')
    ws['A1'] = '各交费期间对应承保年龄范围（标准体与优选体统一适用）'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    hdrs = ['交费期间（年）', '最小承保年龄（岁）', '最大承保年龄（岁）', '备注']
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(2, ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
    ws.row_dimensions[2].height = 22

    age_data = [
        (1,  0, 69, '承保年龄 0~69 岁'),
        (3,  0, 65, '承保年龄 0~65 岁'),
        (5,  0, 62, '承保年龄 0~62 岁'),
        (10, 0, 62, '承保年龄 0~62 岁'),
        (15, 0, 60, '承保年龄 0~60 岁'),
        (20, 0, 59, '承保年龄 0~59 岁'),
        (30, 0, 50, '承保年龄 0~50 岁'),
    ]
    alt = [PatternFill(fill_type='solid', fgColor='DEEAF1'),
           PatternFill(fill_type='solid', fgColor='FFFFFF')]
    for ri, row in enumerate(age_data, 3):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci, value=v)
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = alt[ri % 2]
            cell.alignment = center; cell.border = border
        ws.row_dimensions[ri].height = 22

    # 注：优选体18岁以下不可投保
    ws.merge_cells('A11:D11')
    ws['A11'] = '注：优选体（ensurePlan=2）要求被保人年龄 18~65 周岁，0~17 岁及 66 岁以上仅适用标准体'
    ws['A11'].font = Font(name='微软雅黑', size=9, color='C00000', bold=True)
    ws['A11'].alignment = center
    ws.row_dimensions[11].height = 22

    for ci, w in enumerate([14, 16, 16, 36], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main():
    src_wb = openpyxl.load_workbook(SRC, data_only=True)
    out_wb = openpyxl.Workbook()

    # 提取两个 Sheet 的数据，合并
    all_rows = []
    for sheet_name, plan_code, plan_name in SHEET_MAP:
        ws = src_wb[sheet_name]
        rows = extract_boundary(ws)
        for r in rows:
            r['ensurePlan'] = plan_code
            r['保障方案'] = plan_name
        all_rows.extend(rows)
        print(f'  {sheet_name}: 提取 {len(rows)} 条')

    # 按 ensurePlan → 责任计划 → 交费期间 → 性别 排序
    all_rows.sort(key=lambda x: (x['ensurePlan'], x['责任计划'], x['交费期间'], x['性别']))

    write_summary_sheet(out_wb, all_rows)
    write_age_range_sheet(out_wb)

    out_wb.save(OUT)
    print(f'\n文件已保存：{OUT}')
    print(f'总计：{len(all_rows)} 条边界记录')


if __name__ == '__main__':
    main()
