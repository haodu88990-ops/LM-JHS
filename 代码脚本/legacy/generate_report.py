
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import json

# 读取提取结果
with open(r'f:\workbuddy\test_Jhs\boundary_rates.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

min_age = data['min_age']
max_age = data['max_age']
results = data['results']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '边界费率汇总'

# ====== 样式定义 ======
# 标题样式
title_font = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_fill = PatternFill(fill_type='solid', fgColor='1F4E79')
title_align = Alignment(horizontal='center', vertical='center')

# 表头样式
header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill(fill_type='solid', fgColor='2E75B6')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 子表头（计划/性别分组标题）
sub_header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
sub_fill_plan = [
    PatternFill(fill_type='solid', fgColor='4472C4'),  # 计划0
    PatternFill(fill_type='solid', fgColor='ED7D31'),  # 计划1
    PatternFill(fill_type='solid', fgColor='A9D18E'),  # 计划2（深绿字）
    PatternFill(fill_type='solid', fgColor='FF0000'),  # 计划3
    PatternFill(fill_type='solid', fgColor='7030A0'),  # 计划4
    PatternFill(fill_type='solid', fgColor='00B0F0'),  # 计划5
    PatternFill(fill_type='solid', fgColor='FF6600'),  # 计划6
    PatternFill(fill_type='solid', fgColor='70AD47'),  # 计划7
]
plan_colors = ['4472C4','ED7D31','375623','C00000','7030A0','0070C0','FF6600','375623']

# 行数据样式
data_align = Alignment(horizontal='center', vertical='center')
data_font = Font(name='微软雅黑', size=10)
border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)

# 交替行填充
fill_even = PatternFill(fill_type='solid', fgColor='EBF3FB')
fill_odd = PatternFill(fill_type='solid', fgColor='FFFFFF')

# None 值特殊标注
na_font = Font(name='微软雅黑', size=10, color='9E9E9E', italic=True)
na_fill = PatternFill(fill_type='solid', fgColor='F2F2F2')

# ====== 写入标题 ======
ws.merge_cells('A1:G1')
ws['A1'] = '瑞泰鸿利传世（致享版）终身寿险 — 边界年龄费率汇总表'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = title_align
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:G2')
ws['A2'] = f'保险期间：终身 | 年龄边界：{min_age}岁 ~ {max_age}岁 | 金额单位：元（每1,000元保险金额）'
ws['A2'].font = Font(name='微软雅黑', size=10, italic=True, color='595959')
ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 22

# ====== 写入表头 ======
headers = ['责任计划', '保险期间', '交费期间（年）', '性别', f'{min_age}岁费率', f'{max_age}岁费率', '备注']
for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border
ws.row_dimensions[3].height = 30

# ====== 写入数据 ======
plan_color_map = {
    0: 'DDEEFF', 1: 'FFF3E8', 2: 'F0FFF0', 3: 'FFF0F0',
    4: 'F5F0FF', 5: 'E8F8FF', 6: 'FFF8E8', 7: 'F0F8E8'
}
plan_header_colors = {
    0: '4472C4', 1: 'ED7D31', 2: '375623', 3: 'C00000',
    4: '7030A0', 5: '0070C0', 6: 'FF6600', 7: '375623'
}

data_row = 4
prev_plan = None
for item in results:
    plan = item['责任计划']
    ins = item['保险期间']
    pay = item['交费期间']
    gender = item['性别']
    min_rate = item.get(f'{min_age}岁费率')
    max_rate = item.get(f'{max_age}岁费率')
    
    # 确定行底色
    bg_color = plan_color_map.get(plan, 'FFFFFF')
    row_fill = PatternFill(fill_type='solid', fgColor=bg_color)
    
    values = [plan, ins, pay, gender, min_rate, max_rate]
    remarks = []
    if min_rate is None:
        remarks.append(f'{min_age}岁无数据')
    if max_rate is None:
        remarks.append(f'{max_age}岁无数据')
    values.append('、'.join(remarks) if remarks else '')
    
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=data_row, column=col_idx, value=v)
        cell.alignment = data_align
        cell.border = border
        
        if col_idx in (5, 6) and v is None:
            cell.value = '—'
            cell.font = na_font
            cell.fill = PatternFill(fill_type='solid', fgColor='F2F2F2')
        else:
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = row_fill
            
        # 计划列加粗显示
        if col_idx == 1:
            cell.font = Font(name='微软雅黑', size=10, bold=True, 
                           color=plan_header_colors.get(plan, '000000'))
        # 费率列数字格式
        if col_idx in (5, 6) and v is not None:
            cell.number_format = '0.0'
    
    ws.row_dimensions[data_row].height = 20
    data_row += 1

# ====== 设置列宽 ======
col_widths = [10, 10, 14, 8, 12, 12, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ====== 冻结表头 ======
ws.freeze_panes = 'A4'

# ====== 第二个Sheet：按计划分组透视表 ======
ws2 = wb.create_sheet('按计划分组')

# 组织数据：{计划: {交费期间: {性别: (min_rate, max_rate)}}}
from collections import defaultdict
plan_data = defaultdict(lambda: defaultdict(dict))
pay_periods_set = set()
plans_set = set()

for item in results:
    plan = item['责任计划']
    pay = item['交费期间']
    gender = item['性别']
    min_rate = item.get(f'{min_age}岁费率')
    max_rate = item.get(f'{max_age}岁费率')
    plan_data[plan][pay][gender] = (min_rate, max_rate)
    pay_periods_set.add(pay)
    plans_set.add(plan)

pay_periods = sorted(pay_periods_set)
plans = sorted(plans_set)

r = 1
total_cols = 1 + len(pay_periods) * 2
ws2.merge_cells(start_row=r, end_row=r, start_column=1, end_column=total_cols)
ws2.cell(r, 1).value = '各责任计划 × 交费期间 × 性别 × 年龄 费率透视表'
ws2.cell(r, 1).font = title_font
ws2.cell(r, 1).fill = title_fill
ws2.cell(r, 1).alignment = title_align

r = 2
# 注意：此时第1列不在合并区域，可正常写入
ws2.cell(r, 1).value = '责任计划 / 性别'
ws2.cell(r, 1).font = header_font
ws2.cell(r, 1).fill = header_fill
ws2.cell(r, 1).alignment = header_align
ws2.cell(r, 1).border = border

# 交费期间列头（每个交费期间占2列：男/女）
col = 2
pay_col_map = {}
for pay in pay_periods:
    ws2.merge_cells(start_row=r, end_row=r, start_column=col, end_column=col+1)
    ws2.cell(r, col).value = f'交{pay}年'
    ws2.cell(r, col).font = header_font
    ws2.cell(r, col).fill = PatternFill(fill_type='solid', fgColor='2E75B6')
    ws2.cell(r, col).alignment = header_align
    ws2.cell(r, col).border = border
    pay_col_map[pay] = col
    col += 2

r = 3
ws2.cell(r, 1).value = '年龄/性别'
ws2.cell(r, 1).font = header_font
ws2.cell(r, 1).fill = header_fill
ws2.cell(r, 1).alignment = header_align
ws2.cell(r, 1).border = border
col = 2
for pay in pay_periods:
    ws2.cell(r, col).value = '男'
    ws2.cell(r, col).font = header_font
    ws2.cell(r, col).fill = PatternFill(fill_type='solid', fgColor='4472C4')
    ws2.cell(r, col).alignment = header_align
    ws2.cell(r, col).border = border
    ws2.cell(r, col+1).value = '女'
    ws2.cell(r, col+1).font = header_font
    ws2.cell(r, col+1).fill = PatternFill(fill_type='solid', fgColor='C00000')
    ws2.cell(r, col+1).alignment = header_align
    ws2.cell(r, col+1).border = border
    col += 2

# 数据行：每个计划2行（0岁、69岁）
for plan in plans:
    bg = plan_color_map.get(plan, 'FFFFFF')
    row_fill2 = PatternFill(fill_type='solid', fgColor=bg)
    
    for age_label, age_idx in [(f'{min_age}岁', 0), (f'{max_age}岁', 1)]:
        r += 1
        # 行标签
        label = f'计划{plan} - {age_label}'
        ws2.cell(r, 1).value = label
        ws2.cell(r, 1).font = Font(name='微软雅黑', size=10, bold=True,
                                    color=plan_header_colors.get(plan,'000000'))
        ws2.cell(r, 1).fill = row_fill2
        ws2.cell(r, 1).alignment = data_align
        ws2.cell(r, 1).border = border
        
        # 填充费率
        for pay in pay_periods:
            col = pay_col_map[pay]
            for gi, gender in enumerate(['男', '女']):
                rate_tuple = plan_data[plan][pay].get(gender)
                if rate_tuple:
                    val = rate_tuple[age_idx]
                else:
                    val = None
                cell = ws2.cell(r, col+gi)
                cell.value = val if val is not None else '—'
                cell.font = Font(name='微软雅黑', size=10) if val is not None else na_font
                cell.fill = row_fill2 if val is not None else PatternFill(fill_type='solid', fgColor='F2F2F2')
                cell.alignment = data_align
                cell.border = border
                if val is not None:
                    cell.number_format = '0.0'
        
        ws2.row_dimensions[r].height = 20

# 设置列宽
ws2.column_dimensions['A'].width = 20
for i in range(2, 2+len(pay_periods)*2):
    ws2.column_dimensions[get_column_letter(i)].width = 9
ws2.freeze_panes = 'B4'
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 24
ws2.row_dimensions[3].height = 24

# 保存
out_path = r'f:\workbuddy\test_Jhs\费率边界值汇总.xlsx'
wb.save(out_path)
print(f"已保存到: {out_path}")
print(f"共提取 {len(results)} 条记录")
print(f"其中69岁有费率数据: {sum(1 for r in results if r.get(f'{max_age}岁费率') is not None)} 条")
print(f"其中69岁无费率数据（显示'—'）: {sum(1 for r in results if r.get(f'{max_age}岁费率') is None)} 条")
