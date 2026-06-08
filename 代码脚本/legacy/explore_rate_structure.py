#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
探索费率表.xlsx的结构
"""

import openpyxl

def explore_rate_file():
    """探索费率表结构"""
    
    rate_file = r'f:\workbuddy\test_Jhs\费率表.xlsx'
    wb = openpyxl.load_workbook(rate_file, data_only=True)
    
    print(f"文件: {rate_file}")
    print(f"Sheet列表: {wb.sheetnames}\n")
    
    for sheet_name in wb.sheetnames:
        print(f"=== Sheet: {sheet_name} ===")
        sheet = wb[sheet_name]
        
        # 读取前10行，查看结构
        print("前10行内容:")
        for row in range(1, 11):
            row_data = []
            for col in range(1, min(20, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(cell_value)
            print(f"  行{row}: {row_data[:10]}")  # 只显示前10列
        
        print(f"\n总行数: {sheet.max_row}, 总列数: {sheet.max_column}")
        
        # 查看第4-8行的内容（根据记忆，这些行包含关键信息）
        print("\n关键行内容:")
        for row in range(4, 9):
            row_data = []
            for col in range(1, min(30, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    row_data.append(f"列{col}={cell_value}")
            if row_data:
                print(f"  行{row}: {', '.join(row_data[:15])}")
        
        print("\n" + "="*50 + "\n")

if __name__ == '__main__':
    explore_rate_file()
