#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成中英爱永恒A款分红边界值汇总Excel
根据费率表.xlsx提取边界值费率
"""

import openpyxl
import json
import os

def get_max_age(pay_period):
    """根据交费期间返回最大承保年龄"""
    age_limits = {
        '1': 69,
        '3': 65,
        '5': 62,
        '10': 62,
        '15': 60,
        '20': 59,
        '30': 50
    }
    return age_limits.get(str(pay_period), 69)

def generate_boundary_summary():
    """生成边界值汇总Excel"""
    
    # 读取费率表.xlsx
    rate_file = r'f:\workbuddy\test_Jhs\费率表.xlsx'
    wb = openpyxl.load_workbook(rate_file, data_only=True)
    
    # 获取所有sheet名
    sheet_names = wb.sheetnames
    print(f"找到Sheets: {sheet_names}")
    
    # 准备汇总数据
    summary_data = []
    
    # 处理每个Sheet（标准体费率表和优选体费率表）
    for sheet_name in sheet_names:
        if '费率表' not in sheet_name:
            continue
            
        sheet = wb[sheet_name]
        print(f"\n处理Sheet: {sheet_name}")
        
        # 判断是标准体还是优选体
        ensure_plan = '1' if '标准体' in sheet_name else '2'
        ensure_plan_name = '标准体' if ensure_plan == '1' else '优选体'
        
        # 读取表头信息
        # 行4=计划(0-7), 行5=保险期间, 行6=交费期间, 行7=性别, 行8-77=年龄数据
        
        # 获取所有计划（行4）
        plans = []
        for col in range(2, sheet.max_column + 1):
            plan_value = sheet.cell(row=4, column=col).value
            if plan_value is not None:
                plans.append((col, plan_value))
        
        # 获取所有交费期间（行6）
        pay_periods = []
        for col in range(2, sheet.max_column + 1):
            pay_period = sheet.cell(row=6, column=col).value
            if pay_period is not None:
                pay_periods.append((col, pay_period))
        
        # 去重
        plans = list(set(plans))
        pay_periods = list(set(pay_periods))
        
        print(f"  找到 {len(plans)} 个计划, {len(pay_periods)} 个交费期间")
        
        # 遍历每个交费期间
        for pay_col, pay_period in pay_periods:
            pay_period_str = str(int(pay_period)) if isinstance(pay_period, (int, float)) else str(pay_period)
            max_age = get_max_age(pay_period_str)
            
            # 优选体年龄限制18~65岁
            if ensure_plan == '2':
                min_age = 18
                if max_age > 65:
                    max_age = 65
            else:
                min_age = 0
            
            # 遍历每个计划
            for plan_col, plan in plans:
                if plan_col != pay_col:
                    continue
                    
                plan_int = int(plan) if isinstance(plan, (int, float)) else int(float(plan))
                
                # 读取性别（行7）
                gender_col = plan_col
                gender_value = sheet.cell(row=7, column=gender_col).value
                
                # 遍历性别（1=男, 2=女）
                for gender_code in [1, 2]:
                    # 查找对应的列
                    actual_col = None
                    for col in range(2, sheet.max_column + 1):
                        plan_val = sheet.cell(row=4, column=col).value
                        pay_val = sheet.cell(row=6, column=col).value
                        gender_val = sheet.cell(row=7, column=col).value
                        
                        if plan_val == plan and pay_val == pay_period and gender_val == gender_code:
                            actual_col = col
                            break
                    
                    if actual_col is None:
                        continue
                    
                    # 读取年龄数据（行8-77）
                    age_data = []
                    for row in range(8, 78):
                        age = sheet.cell(row=row, column=1).value
                        rate = sheet.cell(row=row, column=actual_col).value
                        
                        if age is not None and rate is not None:
                            age_int = int(age) if isinstance(age, (int, float)) else int(float(age))
                            rate_float = float(rate) if isinstance(rate, (int, float)) else float(rate)
                            age_data.append((age_int, rate_float))
                    
                    if not age_data:
                        continue
                    
                    # 筛选有效年龄范围内的数据
                    valid_data = [(age, rate) for age, rate in age_data if min_age <= age <= max_age]
                    
                    if not valid_data:
                        continue
                    
                    # 获取最小年龄和最大年龄的费率
                    min_age_data = min(valid_data, key=lambda x: x[0])
                    max_age_data = max(valid_data, key=lambda x: x[0])
                    
                    # 添加到汇总数据
                    summary_data.append({
                        '保障方案': ensure_plan_name,
                        'ensurePlan': ensure_plan,
                        '责任计划': plan_int,
                        '保险期间': '终身',
                        '交费期间(年)': pay_period_str,
                        '性别': '男' if gender_code == 1 else '女',
                        '最小年龄': min_age_data[0],
                        '最小年龄费率': min_age_data[1],
                        '最大年龄': max_age_data[0],
                        '最大年龄费率': max_age_data[1]
                    })
    
    # 创建汇总Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb_new = Workbook()
    ws = wb_new.active
    ws.title = '边界费率汇总'
    
    # 写入表头
    headers = ['保障方案', 'ensurePlan', '责任计划', '保险期间', '交费期间(年)', '性别', '最小年龄', '最小年龄费率', '最大年龄', '最大年龄费率']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # 写入数据
    for row_idx, data in enumerate(summary_data, 2):
        ws.cell(row=row_idx, column=1, value=data['保障方案'])
        ws.cell(row=row_idx, column=2, value=data['ensurePlan'])
        ws.cell(row=row_idx, column=3, value=data['责任计划'])
        ws.cell(row=row_idx, column=4, value=data['保险期间'])
        ws.cell(row=row_idx, column=5, value=data['交费期间(年)'])
        ws.cell(row=row_idx, column=6, value=data['性别'])
        ws.cell(row=row_idx, column=7, value=data['最小年龄'])
        ws.cell(row=row_idx, column=8, value=data['最小年龄费率'])
        ws.cell(row=row_idx, column=9, value=data['最大年龄'])
        ws.cell(row=row_idx, column=10, value=data['最大年龄费率'])
    
    # 调整列宽
    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # 保存文件
    output_file = r'f:\workbuddy\test_Jhs\中英爱永恒A款分红边界值汇总.xlsx'
    wb_new.save(output_file)
    print(f"\n✅ 边界值汇总已生成: {output_file}")
    print(f"   共 {len(summary_data)} 行数据")
    
    return output_file, len(summary_data)

if __name__ == '__main__':
    print("开始生成中英爱永恒A款分红边界值汇总...")
    output_file, row_count = generate_boundary_summary()
    print(f"\n完成！生成了 {row_count} 行边界值数据")
