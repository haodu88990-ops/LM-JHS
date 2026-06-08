
import openpyxl

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
print("所有Sheet:", wb.sheetnames)

ws_main = wb['标准体费率表']
print(f"\n标准体费率表前7行完整内容(前30列):")
for r in range(1, 8):
    row_vals = []
    for c in range(1, 31):
        v = ws_main.cell(row=r, column=c).value
        row_vals.append(str(v) if v is not None else '_')
    print(f"  行{r}: {row_vals}")

# 检查注释
print("\n检查全表注释...")
for r in range(1, ws_main.max_row+1):
    for c in range(1, ws_main.max_column+1):
        cell = ws_main.cell(row=r, column=c)
        if cell.comment:
            print(f"  ({r},{c}): 注释={cell.comment.text}")
