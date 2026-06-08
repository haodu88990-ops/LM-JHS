
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

wb_src = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
ws = wb_src['标准体费率表']

# ====== 读取表头 ======
row4 = {c: ws.cell(row=4, column=c).value for c in range(2, 114)}
row5 = {c: ws.cell(row=5, column=c).value for c in range(2, 114)}
row6 = {c: ws.cell(row=6, column=c).value for c in range(2, 114)}
row7 = {c: str(ws.cell(row=7, column=c).value).strip() for c in range(2, 114)}

# 年龄行映射
age_rows = {}
for r in range(8, 78):
    age = ws.cell(row=r, column=1).value
    if age is not None:
        age_rows[age] = r
ages = sorted(age_rows.keys())

# 列映射
col_map = {}
for c in range(2, 114):
    plan = row4.get(c)
    period = row5.get(c)
    pay = row6.get(c)
    gender = row7.get(c, '').strip()
    if plan is not None and period and pay and gender in ('男', '女'):
        col_map[c] = {'计划': int(plan), '保险期间': str(period), '交费期间': int(pay), '性别': gender}

# ====== 动态提取每列的最小/最大有效年龄及其费率 ======
results = []
for c in sorted(col_map.keys()):
    info = col_map[c]
    min_age_val = max_age_val = min_rate = max_rate = None
    for age in ages:
        val = ws.cell(row=age_rows[age], column=c).value
        if val is not None and val != '':
            if min_age_val is None:
                min_age_val = age
                min_rate = val
            max_age_val = age
            max_rate = val
    results.append({
        '责任计划': info['计划'],
        '保险期间': info['保险期间'],
        '交费期间': info['交费期间'],
        '性别': info['性别'],
        '最小年龄': min_age_val,
        '最大年龄': max_age_val,
        '最小年龄费率': min_rate,
        '最大年龄费率': max_rate,
    })

# ====== 样式定义 ======
title_font    = Font(name='微软雅黑', size=14, bold=True, color='FFFFFF')
title_fill    = PatternFill(fill_type='solid', fgColor='1F4E79')
header_font   = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
header_fill   = PatternFill(fill_type='solid', fgColor='2E75B6')
sub_fill      = PatternFill(fill_type='solid', fgColor='4472C4')
center        = Alignment(horizontal='center', vertical='center', wrap_text=True)
na_font       = Font(name='微软雅黑', size=10, color='9E9E9E', italic=True)
thin          = Side(style='thin', color='BFBFBF')
border        = Border(left=thin, right=thin, top=thin, bottom=thin)

plan_bg = {0:'DDEEFF', 1:'FFF3E8', 2:'EAFFF0', 3:'FFF0F0',
           4:'F5F0FF', 5:'E8F8FF', 6:'FFF8E8', 7:'F0F8E8'}
plan_fc = {0:'1F4E79', 1:'843C0C', 2:'375623', 3:'C00000',
           4:'7030A0', 5:'0070C0', 6:'C55A11', 7:'375623'}

# ====== 创建工作簿 ======
wb_out = openpyxl.Workbook()

# ──────────────────────────────────────────
# Sheet1：明细汇总（动态年龄边界）
# ──────────────────────────────────────────
ws1 = wb_out.active
ws1.title = '边界费率汇总'

# 标题
ws1.merge_cells('A1:H1')
ws1['A1'] = '瑞泰鸿利传世（致享版）终身寿险 — 各交费期间年龄边界费率汇总表'
ws1['A1'].font = title_font
ws1['A1'].fill = title_fill
ws1['A1'].alignment = center
ws1.row_dimensions[1].height = 36

ws1.merge_cells('A2:H2')
ws1['A2'] = '保险期间：终身 | 各交费期间对应的最小/最大承保年龄不同，本表自动提取对应费率 | 金额单位：元（每1,000元保险金额）'
ws1['A2'].font = Font(name='微软雅黑', size=9, italic=True, color='595959')
ws1['A2'].alignment = center
ws1.row_dimensions[2].height = 20

# 表头
headers = ['责任计划', '保险期间', '交费期间（年）', '性别', '最小年龄（岁）', '最小年龄费率', '最大年龄（岁）', '最大年龄费率']
for ci, h in enumerate(headers, 1):
    cell = ws1.cell(row=3, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws1.row_dimensions[3].height = 30

# 数据
for ri, item in enumerate(results, 4):
    plan = item['责任计划']
    bg = PatternFill(fill_type='solid', fgColor=plan_bg.get(plan, 'FFFFFF'))
    row_vals = [
        item['责任计划'], item['保险期间'], item['交费期间'], item['性别'],
        item['最小年龄'], item['最小年龄费率'],
        item['最大年龄'], item['最大年龄费率'],
    ]
    for ci, v in enumerate(row_vals, 1):
        cell = ws1.cell(row=ri, column=ci, value=v)
        cell.alignment = center
        cell.border = border
        cell.fill = bg
        if ci == 1:
            cell.font = Font(name='微软雅黑', size=10, bold=True, color=plan_fc.get(plan, '000000'))
        elif v is None:
            cell.value = '—'
            cell.font = na_font
        else:
            cell.font = Font(name='微软雅黑', size=10)
        if ci in (6, 8) and v is not None:
            cell.number_format = '0.0'
    ws1.row_dimensions[ri].height = 20

# 列宽
for ci, w in enumerate([10, 10, 14, 8, 14, 14, 14, 14], 1):
    ws1.column_dimensions[get_column_letter(ci)].width = w
ws1.freeze_panes = 'A4'

# ──────────────────────────────────────────
# Sheet2：透视表（行=计划+性别+年龄边界，列=交费期间）
# ──────────────────────────────────────────
ws2 = wb_out.create_sheet('透视表（计划×交费期间）')

pay_periods = sorted(set(r['交费期间'] for r in results))

# 预构建索引
data_idx = {}
for item in results:
    key = (item['责任计划'], item['性别'], item['交费期间'])
    data_idx[key] = item

plans = sorted(set(r['责任计划'] for r in results))

# 标题行
total_cols = 1 + len(pay_periods) * 2
ws2.merge_cells(start_row=1, end_row=1, start_column=1, end_column=total_cols)
ws2.cell(1, 1).value = '各责任计划 × 交费期间 × 性别 费率透视表（自动识别年龄边界）'
ws2.cell(1, 1).font = title_font
ws2.cell(1, 1).fill = title_fill
ws2.cell(1, 1).alignment = center
ws2.row_dimensions[1].height = 32

# 交费期间列头（第2行）
ws2.cell(2, 1).value = '责任计划 / 条件'
ws2.cell(2, 1).font = header_font
ws2.cell(2, 1).fill = header_fill
ws2.cell(2, 1).alignment = center
ws2.cell(2, 1).border = border

col = 2
pay_col_map = {}
for pay in pay_periods:
    ws2.merge_cells(start_row=2, end_row=2, start_column=col, end_column=col+1)
    cell = ws2.cell(2, col)
    cell.value = f'交费{pay}年'
    cell.font = header_font
    cell.fill = PatternFill(fill_type='solid', fgColor='2E75B6')
    cell.alignment = center
    cell.border = border
    pay_col_map[pay] = col
    col += 2

# 性别子表头（第3行）
ws2.cell(3, 1).value = '性别 / 年龄边界'
ws2.cell(3, 1).font = header_font
ws2.cell(3, 1).fill = header_fill
ws2.cell(3, 1).alignment = center
ws2.cell(3, 1).border = border

col = 2
for pay in pay_periods:
    ws2.cell(3, col).value = '男'
    ws2.cell(3, col).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    ws2.cell(3, col).fill = PatternFill(fill_type='solid', fgColor='4472C4')
    ws2.cell(3, col).alignment = center
    ws2.cell(3, col).border = border
    ws2.cell(3, col+1).value = '女'
    ws2.cell(3, col+1).font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    ws2.cell(3, col+1).fill = PatternFill(fill_type='solid', fgColor='C00000')
    ws2.cell(3, col+1).alignment = center
    ws2.cell(3, col+1).border = border
    col += 2

# 数据行：每个计划 × 4行（最小年龄：男，最小年龄：女，最大年龄：男，最大年龄：女）
# 更紧凑：每个计划 2行（最小年龄/最大年龄），每列写男女
# 重新设计：每个计划 2行（行1=最小年龄费率，行2=最大年龄费率），每列内：男 / 女 各1格

current_row = 4
for plan in plans:
    bg = plan_bg.get(plan, 'FFFFFF')
    fc = plan_fc.get(plan, '000000')
    row_fill = PatternFill(fill_type='solid', fgColor=bg)
    
    for age_label, age_key, rate_key in [
        ('最小年龄费率', '最小年龄', '最小年龄费率'),
        ('最大年龄费率', '最大年龄', '最大年龄费率'),
    ]:
        # 行标签
        ws2.cell(current_row, 1).value = f'计划{plan}  {age_label}'
        ws2.cell(current_row, 1).font = Font(name='微软雅黑', size=10, bold=True, color=fc)
        ws2.cell(current_row, 1).fill = row_fill
        ws2.cell(current_row, 1).alignment = center
        ws2.cell(current_row, 1).border = border

        for pay in pay_periods:
            col = pay_col_map[pay]
            for gi, gender in enumerate(['男', '女']):
                item = data_idx.get((plan, gender, pay))
                if item:
                    age_v = item[age_key]
                    rate_v = item[rate_key]
                    display = f'{rate_v}' if rate_v is not None else '—'
                    # 在括号里标注年龄
                    if age_v is not None and rate_v is not None:
                        display = f'{rate_v}\n({age_v}岁)'
                else:
                    display = '—'
                cell = ws2.cell(current_row, col+gi, value=display)
                cell.font = Font(name='微软雅黑', size=9) if display != '—' else na_font
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
        
        ws2.row_dimensions[current_row].height = 30
        current_row += 1

# 列宽
ws2.column_dimensions['A'].width = 22
for i in range(2, 2 + len(pay_periods)*2):
    ws2.column_dimensions[get_column_letter(i)].width = 11
ws2.freeze_panes = 'B4'
ws2.row_dimensions[2].height = 24
ws2.row_dimensions[3].height = 24

# ──────────────────────────────────────────
# Sheet3：按交费期间分组说明
# ──────────────────────────────────────────
ws3 = wb_out.create_sheet('承保年龄范围说明')

ws3.merge_cells('A1:D1')
ws3.cell(1,1).value = '各交费期间 对应承保年龄范围（统一适用全部计划）'
ws3.cell(1,1).font = title_font
ws3.cell(1,1).fill = title_fill
ws3.cell(1,1).alignment = center
ws3.row_dimensions[1].height = 32

for ci, h in enumerate(['交费期间（年）', '最小承保年龄（岁）', '最大承保年龄（岁）', '备注'], 1):
    cell = ws3.cell(2, ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border
ws3.row_dimensions[2].height = 26

# 统计各交费期间的年龄边界（从results取）
pay_age_range = {}
for item in results:
    pay = item['交费期间']
    if pay not in pay_age_range:
        pay_age_range[pay] = {'min': item['最小年龄'], 'max': item['最大年龄']}

row_colors = ['EBF3FB', 'FFFFFF']
for ri, pay in enumerate(sorted(pay_age_range.keys()), 3):
    rng = pay_age_range[pay]
    bg = PatternFill(fill_type='solid', fgColor=row_colors[ri % 2])
    vals = [pay, rng['min'], rng['max'], f"承保年龄 {rng['min']}~{rng['max']} 岁，超龄无费率"]
    for ci, v in enumerate(vals, 1):
        cell = ws3.cell(ri, ci, value=v)
        cell.font = Font(name='微软雅黑', size=10)
        cell.fill = bg
        cell.alignment = center
        cell.border = border
    ws3.row_dimensions[ri].height = 22

for ci, w in enumerate([16, 20, 20, 32], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

# ====== 保存 ======
out_path = r'f:\workbuddy\test_Jhs\费率边界值汇总.xlsx'
wb_out.save(out_path)
print(f"保存成功: {out_path}")
print(f"共 {len(results)} 条记录")
print("\n各交费期间最大承保年龄：")
for pay in sorted(pay_age_range.keys()):
    print(f"  交费{pay}年: {pay_age_range[pay]['min']}岁 ~ {pay_age_range[pay]['max']}岁")
print("\n示例验证（计划1，交5年，男）:")
sample = data_idx.get((1, '男', 5))
if sample:
    print(f"  最小年龄={sample['最小年龄']}岁，费率={sample['最小年龄费率']}")
    print(f"  最大年龄={sample['最大年龄']}岁，费率={sample['最大年龄费率']}")
