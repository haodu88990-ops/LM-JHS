
import openpyxl

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
ws = wb['标准体费率表']

print(f"总行数: {ws.max_row}, 总列数: {ws.max_column}")
print("\n=== 前10行全部列 ===")
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    print(list(row))

print("\n=== 第4行（计划行）所有值 ===")
row4 = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column+1)]
print(row4)

print("\n=== 第5行（保险期间行）所有值 ===")
row5 = [ws.cell(row=5, column=c).value for c in range(1, ws.max_column+1)]
print(row5)

print("\n=== 第6行（交费期间行）所有值 ===")
row6 = [ws.cell(row=6, column=c).value for c in range(1, ws.max_column+1)]
print(row6)

print("\n=== 第7行（性别行）所有值 ===")
row7 = [ws.cell(row=7, column=c).value for c in range(1, ws.max_column+1)]
print(row7)

print("\n=== 最后几行(第70-77行) ===")
for row in ws.iter_rows(min_row=70, max_row=77, values_only=True):
    print(list(row))
