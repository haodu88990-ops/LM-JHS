#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成中英爱永恒A款分红边界值汇总Excel
根据费率表.xlsx提取边界值费率
格式参考：瑞泰鸿利致享版边界值汇总.xlsx
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import datetime

def get_max_age(pay_period_str):
    """根据交费期间返回最大承保年龄"""
    if pay_period_str in ['趸交', '1年']:
        return 69
    elif pay_period_str == '3年':
        return 65
    elif pay_period_str == '5年':
        return 62
    elif pay_period_str in ['10年', '15年']:
        return 62
    elif pay_period_str == '20年':
        return 59
    elif pay_period_str == '30年':
        return 50
    else:
        return 69

def generate_boundary_summary():
    """生成边界值汇总Excel"""
    
    # 读取费率表.xlsx
    rate_file = r'f:\workbuddy\test_Jhs\费率表.xlsx'
    wb = openpyxl.load_workbook(rate_file, data_only=True)
    ws = wb['基本保险金额对应费率表']
    
    print(f"读取文件: {rate_file}")
    print(f"Sheet: 基本保险金额对应费率表")
    print(f"数据范围: 第10行到第{ws.max_row}行")
    
    # 解析表头信息（第6-9行）
    # 第6行: 保障方案（方案一、方案二）
    # 第7行: 保险期间（终身）
    # 第8行: 交费期间（趸交、3年、5年、10年、15年、20年、30年）
    # 第9行: 性别（男、女）
    
    # 收集所有列的信息
    columns_info = []
    for col in range(2, ws.max_column + 1):
        plan = ws.cell(row=6, column=col).value  # 方案一 or 方案二
        period = ws.cell(row=7, column=col).value  # 终身
        pay_period = ws.cell(row=8, column=col).value  # 趸交/3年/5年...
        gender = ws.cell(row=9, column=col).value  # 男/女
        
        if plan and pay_period and gender and period:
            columns_info.append({
                'col': col,
                'plan': str(plan).strip(),
                'period': str(period).strip(),
                'pay_period': str(pay_period).strip(),
                'gender': str(gender).strip()
            })
    
    print(f"\n找到 {len(columns_info)} 个数据列")
    
    # 按保障方案和交费期间分组
    groups = {}
    for info in columns_info:
        key = (info['plan'], info['pay_period'])
        if key not in groups:
            groups[key] = []
        groups[key].append(info)
    
    print(f"分组数: {len(groups)}")
    for key in sorted(groups.keys()):
        print(f"  {key}: {len(groups[key])} 列")
    
    # 准备汇总数据
    summary_data = []
    
    # 遍历每个分组
    for (plan, pay_period), info_list in sorted(groups.items()):
        print(f"\n处理: {plan} - {pay_period}")
        
        max_age = get_max_age(pay_period)
        min_age = 0  # 标准体最小年龄0岁
        
        print(f"  年龄范围: {min_age} ~ {max_age} 岁")
        
        # 遍历该分组的所有列（男、女）
        for info in info_list:
            col = info['col']
            gender = info['gender']
            
            # 读取该列的所有年龄和费率（第10行开始）
            age_rate_data = []
            for row in range(10, ws.max_row + 1):
                age = ws.cell(row=row, column=1).value  # 第1列是年龄
                rate = ws.cell(row=row, column=col).value  # 该列的费率
                
                if age is not None and rate is not None and str(rate).strip() != '':
                    try:
                        age_int = int(float(str(age).strip()))
                        rate_float = float(str(rate).strip())
                        
                        # 只保留有效年龄范围内的数据
                        if min_age <= age_int <= max_age:
                            age_rate_data.append((age_int, rate_float))
                    except (ValueError, TypeError):
                        continue
            
            if not age_rate_data:
                print(f"  {gender}: 无有效数据")
                continue
            
            # 找出最小年龄和最大年龄的费率
            min_age_data = min(age_rate_data, key=lambda x: x[0])
            max_age_data = max(age_rate_data, key=lambda x: x[0])
            
            # 添加到汇总数据
            summary_data.append({
                '保障方案': plan,
                'ensurePlan': '1',  # 标准体
                '责任计划': 0,  # 默认0，根据实际需要调整
                '保险期间': '终身',
                '交费期间(年)': pay_period,
                '性别': gender,
                '最小年龄': min_age_data[0],
                '最小年龄费率': min_age_data[1],
                '最大年龄': max_age_data[0],
                '最大年龄费率': max_age_data[1]
            })
            
            print(f"  {gender}: 最小年龄={min_age_data[0]}, 最大年龄={max_age_data[0]}, 费率={min_age_data[1]}/{max_age_data[1]}")
    
    # 创建汇总Excel
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = '边界费率汇总'
    
    # 添加标题行（合并单元格）
    ws_new.merge_cells('A1:J1')
    title_cell = ws_new.cell(row=1, column=1, value='中英爱永恒终身寿险A款（分红型）— 边界费率汇总表')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加说明行
    ws_new.merge_cells('A2:J2')
    desc_cell = ws_new.cell(row=2, column=1, value='保险期间：终身 | ensurePlan: 1=标准体 | 金额单位：元（每1,000元保额）')
    desc_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入表头（第3行）
    headers = ['保障方案', 'ensurePlan', '责任计划', '保险期间', 
               '交费期间（年）', '性别', '最小年龄（岁）', '最小年龄费率', 
               '最大年龄（岁）', '最大年龄费率']
    
    for col, header in enumerate(headers, 1):
        cell = ws_new.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    
    # 写入数据（第4行开始）
    for row_idx, data in enumerate(summary_data, 4):
        ws_new.cell(row=row_idx, column=1, value=data['保障方案'])
        ws_new.cell(row=row_idx, column=2, value=data['ensurePlan'])
        ws_new.cell(row=row_idx, column=3, value=data['责任计划'])
        ws_new.cell(row=row_idx, column=4, value=data['保险期间'])
        ws_new.cell(row=row_idx, column=5, value=data['交费期间(年)'])
        ws_new.cell(row=row_idx, column=6, value=data['性别'])
        ws_new.cell(row=row_idx, column=7, value=data['最小年龄'])
        ws_new.cell(row=row_idx, column=8, value=data['最小年龄费率'])
        ws_new.cell(row=row_idx, column=9, value=data['最大年龄'])
        ws_new.cell(row=row_idx, column=10, value=data['最大年龄费率'])
    
    # 调整列宽
    column_widths = [15, 12, 12, 12, 18, 8, 15, 15, 15, 15]
    for col in range(1, 11):
        ws_new.column_dimensions[openpyxl.utils.get_column_letter(col)].width = column_widths[col-1]
    
    # 添加生成时间
    ws_new['A' + str(len(summary_data) + 5)] = f"生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # 保存文件
    output_file = r'f:\workbuddy\test_Jhs\中英爱永恒A款分红边界值汇总.xlsx'
    wb_new.save(output_file)
    
    print(f"\n✅ 边界值汇总已生成: {output_file}")
    print(f"   共 {len(summary_data)} 行数据")
    
    return output_file, len(summary_data)

if __name__ == '__main__':
    print("="*70)
    print("开始生成中英爱永恒A款分红边界值汇总...")
    print("="*70)
    
    try:
        output_file, row_count = generate_boundary_summary()
        
        print("\n" + "="*70)
        print(f"✅ 完成！生成了 {row_count} 行边界值数据")
        print(f"   文件保存在: {output_file}")
        print("="*70)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
