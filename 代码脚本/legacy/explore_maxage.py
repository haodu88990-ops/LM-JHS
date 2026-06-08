
import openpyxl

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
ws = wb['标准体费率表']

# 读取表头行
row4 = {c: ws.cell(row=4, column=c).value for c in range(2, 114)}  # 责任计划
row5 = {c: ws.cell(row=5, column=c).value for c in range(2, 114)}  # 保险期间
row6 = {c: ws.cell(row=6, column=c).value for c in range(2, 114)}  # 交费期间
row7 = {c: str(ws.cell(row=7, column=c).value).strip() for c in range(2, 114)}  # 性别

# 读取所有年龄行
age_rows = {}
for r in range(8, 78):
    age = ws.cell(row=r, column=1).value
    if age is not None:
        age_rows[age] = r
ages = sorted(age_rows.keys())

# 建立列映射
col_map = {}
for c in range(2, 114):
    plan = row4.get(c)
    period = row5.get(c)
    pay = row6.get(c)
    gender = row7.get(c, '').strip()
    if plan is not None and period and pay and gender in ('男', '女'):
        col_map[c] = {'计划': int(plan), '保险期间': str(period), '交费期间': int(pay), '性别': gender}

# 对每列找最大有效年龄（最后一个非None非空字符串的行）
print(f"{'列':>4} {'计划':>4} {'交费':>4} {'性别':>4} {'最小年龄':>6} {'最大年龄':>6} {'最小费率':>8} {'最大费率':>8}")
print("-" * 60)

results = []
for c in sorted(col_map.keys()):
    info = col_map[c]
    # 找该列最大有效年龄
    min_age_val = None
    max_age_val = None
    min_rate = None
    max_rate = None
    
    for age in ages:
        row_idx = age_rows[age]
        val = ws.cell(row=row_idx, column=c).value
        # 有效值：非None，且不是空字符串
        if val is not None and val != '':
            if min_age_val is None:
                min_age_val = age
                min_rate = val
            max_age_val = age
            max_rate = val
    
    results.append({
        '责任计划': info['计划'],
        '保险期间': info['保险期间'],
        '交费期间': info['交费期间'],
        '性别': info['性别'],
        '最小年龄': min_age_val,
        '最大年龄': max_age_val,
        '最小年龄费率': min_rate,
        '最大年龄费率': max_rate,
    })
    print(f"{c:>4} {info['计划']:>4} {info['交费期间']:>4} {info['性别']:>4} {str(min_age_val):>6} {str(max_age_val):>6} {str(min_rate):>8} {str(max_rate):>8}")

# 统计不同交费期间对应的最大年龄情况
print("\n=== 各交费期间对应的最大年龄统计 ===")
from collections import defaultdict
pay_max_age = defaultdict(set)
for r in results:
    pay_max_age[r['交费期间']].add(r['最大年龄'])
for pay in sorted(pay_max_age.keys()):
    print(f"  交费{pay}年: 最大年龄 = {sorted(pay_max_age[pay])}")
