
import openpyxl

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
ws = wb['标准体费率表']

print("=== 第4行（计划行）所有非None值及列索引 ===")
row4 = [(c, ws.cell(row=4, column=c).value) for c in range(1, ws.max_column+1)]
for c, v in row4:
    if v is not None and v != '':
        print(f"  列{c}: {v}")

print("\n=== 第5行（保险期间行）所有非None值及列索引 ===")
row5 = [(c, ws.cell(row=5, column=c).value) for c in range(1, ws.max_column+1)]
for c, v in row5:
    if v is not None and v != '':
        print(f"  列{c}: {v}")

print("\n=== 第6行（交费期间行）所有非None值及列索引 ===")
row6 = [(c, ws.cell(row=6, column=c).value) for c in range(1, ws.max_column+1)]
for c, v in row6:
    if v is not None and v != '':
        print(f"  列{c}: {v}")

print("\n=== 第7行（性别行）所有非空值及列索引 ===")
row7 = [(c, ws.cell(row=7, column=c).value) for c in range(1, ws.max_column+1)]
for c, v in row7:
    if v is not None and v != '':
        print(f"  列{c}: {v}")

print("\n=== 第77行（最后一行，应是70岁或末行） ===")
row_last = [(c, ws.cell(row=77, column=c).value) for c in range(1, ws.max_column+1)]
non_empty = [(c,v) for c,v in row_last if v is not None and v != '']
print(non_empty[:30])
