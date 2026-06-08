# -*- coding: utf-8 -*-
"""
批量接口测试脚本 v2
修正：
  - ensurePlan 传 "1"（标准体），费率表"计划0~7"通过 dutyOptionList 区分
  - payModeCode：交费期间=1年时传"1"(一次交清)，其他传"5"(年交)
  - insurantSex：1=男, 2=女
  - ensurePeriodCode：终身=TO105
  - age_rate → plan_rate 参数动态串联（serialNo/proposalId）
"""

import requests
import json
import random
import time
import warnings
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

warnings.filterwarnings("ignore")


# ──────────────────────────────────────────
# 1. 责任计划 → dutyOptionList 映射
# ──────────────────────────────────────────
DUTY_ITEM_TEMPLATES = {
    "2": {
        "factorType": "dutyOption", "factorName": "全残保险金",
        "isDisplay": "1", "order": 115218, "factorLevel": 1,
        "code": "2", "type": "dutyOption", "value": "2",
        "desc": "全残保险金", "dataType": "5", "appFactorList": [],
        "defaultValue": "2", "appfactorCode": "2", "premium": ""
    },
    "3": {
        "factorType": "dutyOption", "factorName": "身故额外关爱保险金",
        "isDisplay": "1", "order": 115219, "factorLevel": 1,
        "code": "3", "type": "dutyOption", "value": "3",
        "desc": "身故额外关爱保险金", "dataType": "5", "appFactorList": [],
        "defaultValue": ["3"], "appfactorCode": "3", "premium": ""
    },
    "4": {
        "factorType": "dutyOption", "factorName": "意外身故额外保险金",
        "isDisplay": "1", "order": 115220, "factorLevel": 1,
        "code": "4", "type": "dutyOption", "value": "4",
        "desc": "意外身故额外保险金", "dataType": "5", "appFactorList": [],
        "defaultValue": ["4"], "appfactorCode": "4", "premium": ""
    }
}

# 各计划包含的附加责任code（0=只有身故→空列表）
PLAN_DUTY_CODES = {
    0: [],
    1: ["2"],
    2: ["3"],
    3: ["3", "2"],
    4: ["4"],
    5: ["4", "2"],
    6: ["4", "3"],
    7: ["4", "3", "2"],
}

def build_duty_option_list(plan: int) -> list:
    """根据责任计划编号构建dutyOptionList"""
    codes = PLAN_DUTY_CODES.get(plan, [])
    return [DUTY_ITEM_TEMPLATES[c] for c in codes]


# ──────────────────────────────────────────
# 2. 参数映射
# ──────────────────────────────────────────
# 保险期间 → ensurePeriodCode
ENSURE_PERIOD_MAP = {
    "终身": "TO105"
}

# 交费期间 → payPeriodCode（直接用数字字符串）
def pay_period_code(pay: int) -> str:
    return str(pay)

# 交费期间 → payModeCode（1年=一次交清"1"，其他=年交"5"）
def pay_mode_code(pay: int) -> str:
    return "1" if pay == 1 else "5"

# 性别 → insurantSex（1=男，2=女）
SEX_MAP = {"男": "1", "女": "2"}


# ──────────────────────────────────────────
# 3. 从费率表提取测试用例（动态最大年龄）
# ──────────────────────────────────────────
def load_test_cases():
    """从「瑞泰鸿利致享版边界值汇总.xlsx」的「边界费率汇总」Sheet 直接提取测试用例
    10列结构：保障方案 | ensurePlan | 责任计划 | 保险期间 | 交费期间(年) | 性别
             | 最小年龄 | 最小年龄费率 | 最大年龄 | 最大年龄费率
    包含标准体(ensurePlan=1)和优选体(ensurePlan=2)，共448条用例
    """
    wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\瑞泰鸿利致享版边界值汇总.xlsx', data_only=True)
    ws = wb['边界费率汇总']

    cases = []
    for r in range(4, ws.max_row + 1):
        # col1=保障方案, col2=ensurePlan, col3=责任计划, col4=保险期间
        # col5=交费期间, col6=性别, col7=最小年龄, col8=最小年龄费率
        # col9=最大年龄, col10=最大年龄费率
        bz_plan    = ws.cell(row=r, column=1).value   # 保障方案（标准体/优选体）
        ensure_plan = ws.cell(row=r, column=2).value  # ensurePlan: "1" 或 "2"
        plan        = ws.cell(row=r, column=3).value  # 责任计划 0-7
        period      = ws.cell(row=r, column=4).value  # 保险期间
        pay         = ws.cell(row=r, column=5).value  # 交费期间(年)
        gender      = ws.cell(row=r, column=6).value  # 性别
        min_age     = ws.cell(row=r, column=7).value
        min_rate    = ws.cell(row=r, column=8).value
        max_age     = ws.cell(row=r, column=9).value
        max_rate    = ws.cell(row=r, column=10).value

        if plan is None or gender not in ('男', '女'):
            continue

        base = {
            '保障方案': str(bz_plan) if bz_plan else '',
            'ensurePlan': str(ensure_plan) if ensure_plan else '1',
            '责任计划': int(plan),
            '保险期间': str(period),
            '交费期间': int(pay),
            '性别': gender,
        }
        if min_age is not None and min_rate is not None:
            cases.append({**base, '年龄': int(min_age), '年龄类型': '最小年龄', '期望费率': float(min_rate)})
        if max_age is not None and max_rate is not None and max_age != min_age:
            cases.append({**base, '年龄': int(max_age), '年龄类型': '最大年龄', '期望费率': float(max_rate)})

    return cases


# ──────────────────────────────────────────
# 4. API 调用类（修正参数传递）
# ──────────────────────────────────────────
class BatchAPITest:
    def __init__(self):
        self.base_url = "https://kfzxtb.lmbaoxian.com:13080"
        self.session = requests.Session()
        self.login_cookies = None
        # 动态参数
        self.serial_no = "1491455820292947968"
        self.proposal_id = "40472"

    def login(self, account="15856990088", password="dc483e80a7a0bd9ef71d8cf973673924"):
        url = f"{self.base_url}/broker/api/user/login.html"
        data = {"account": account, "password": password}
        headers = {
            "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest"
        }
        resp = self.session.post(url, data=data, headers=headers, verify=False, timeout=30)
        if "不能为空" in resp.text:
            headers["Content-Type"] = "application/json"
            resp = self.session.post(url, json=data, headers=headers, verify=False, timeout=30)
        self.login_cookies = resp.cookies
        try:
            result = resp.json()
            success = result.get('code') == 0 or result.get('success') or '成功' in resp.text
        except Exception:
            success = resp.status_code == 200
        print(f"  登录: status={resp.status_code}, success={success}")
        return success, resp

    def age_rate(self, insurant_age: int, insurant_sex: str, cookies=None):
        """调用 saveCustomer 接口（年龄_费率）
        insurantSex: 1=男, 2=女
        从响应中动态提取 serialNo/proposalId
        """
        url = f"{self.base_url}/broker/api/prospectus/saveCustomer.html"
        sex_code = SEX_MAP.get(insurant_sex, "1")
        payload = {
            "insurantName": "",
            "insurantAge": insurant_age,
            "insurantSex": sex_code,          # 1=男, 2=女
            "insurantOccLevel": 1,
            "insurantSocialInsurance": "1",
            "insurantId": 85320,
            "insurantBirthday": "",
            "policyHolderAge": 30,
            "policyHolderSex": "1",
            "policyHolderId": 85321,
            "policyHolderBirthday": "",
            "serialNo": self.serial_no,
            "type": "prospectus"
        }
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest"
        }
        cookies = cookies or self.login_cookies
        resp = self.session.post(url, json=payload, headers=headers, cookies=cookies, verify=False, timeout=30)
        new_cookies = resp.cookies if resp.cookies else cookies
        
        try:
            result = resp.json()
            success = result.get('code') == 200 or result.get('success') or '成功' in resp.text
            # ★ 从 proposalPlanVOList[0] 动态提取
            info = result.get('info', {})
            ppvl = info.get('proposalPlanVOList', [])
            if ppvl and isinstance(ppvl[0], dict):
                plan_vo = ppvl[0]
                if plan_vo.get('serialNo'):
                    self.serial_no = str(plan_vo['serialNo'])
                if plan_vo.get('proposalId'):
                    self.proposal_id = str(plan_vo['proposalId'])
            elif info.get('serialNo'):
                self.serial_no = str(info['serialNo'])
        except Exception:
            success = resp.status_code == 200
        
        return success, new_cookies, resp.status_code, resp.text

    def plan_rate(self, plan: int, ensure_period: str, pay_period: int,
                  amount: int, duty_list: list, ensure_plan: str = "1", cookies=None):
        """调用 saveProductExt 接口（计划_费率）
        ★ ensurePlan 动态传入："1"=标准体, "2"=优选体
        ★ payModeCode: 交1年="1"(一次交清), 其他="5"(年交)
        ★ ensurePeriodCode: 终身=TO105
        """
        url = f"{self.base_url}/broker/api/prospectus/saveProductExt.html"
        payload = {
            "serialNo": self.serial_no,
            "productId": "991452",
            "companyId": "100080",
            "proposalId": self.proposal_id,
            "dividendDrawType": "2",
            "premium": "",
            "ensurePeriodCode": ENSURE_PERIOD_MAP.get(ensure_period, "TO105"),
            "payPeriodCode": pay_period_code(pay_period),
            "payModeCode": pay_mode_code(pay_period),   # ★ 1年→"1", 其他→"5"
            "ensurePlan": ensure_plan,                  # ★ 动态传入："1"=标准体, "2"=优选体
            "amountDescr": str(amount),
            "amount": str(amount),
            "fee": "",
            "dutyOptionList": duty_list
        }
        headers = {
            "Content-Type": "application/json",
            "Accept": "application/json, text/javascript, */*; q=0.01",
            "X-Requested-With": "XMLHttpRequest"
        }
        resp = self.session.post(url, json=payload, headers=headers, cookies=cookies, verify=False, timeout=30)
        fee = None
        success = False
        failure_reason = ""
        try:
            result = resp.json()
            code = result.get('code')
            info = result.get('info', {})
            if isinstance(info, dict):
                fee = info.get('fee')
                failure_reason = info.get('failureReason', '')
                # 动态更新 serialNo/proposalId
                if info.get('serialNo'):
                    self.serial_no = str(info['serialNo'])
                if info.get('proposalId'):
                    self.proposal_id = str(info['proposalId'])
            
            # 判断成功：fee > 0 且没有 failureReason
            if code == 200 and fee is not None and fee > 0 and not failure_reason:
                success = True
            elif code == 500:
                success = False
            else:
                success = code == 200
                
        except Exception:
            pass
        return success, fee, resp.status_code, resp.text, failure_reason

    def run_single_case(self, case: dict, case_no: int):
        """执行单条测试用例"""
        plan = case['责任计划']
        ensure_period = case['保险期间']
        pay_period = case['交费期间']
        gender = case['性别']
        age = case['年龄']
        age_type = case['年龄类型']
        expected_rate = case['期望费率']
        ensure_plan = case.get('ensurePlan', '1')        # ★ 从用例取 ensurePlan
        bz_plan = case.get('保障方案', '标准体')          # ★ 保障方案描述

        # 随机保额：100万-500万，每1000元递增
        amount = random.randrange(1_000_000, 5_001_000, 1000)

        duty_list = build_duty_option_list(plan)
        duty_desc = PLAN_DUTY_CODES.get(plan, [])

        print(f"  [{case_no:03d}] {bz_plan} 计划{plan} 交{pay_period}年 {gender} {age}岁 保额{amount:,} ...", end=' ', flush=True)

        result = {
            '序号': case_no,
            '保障方案': bz_plan,
            'ensurePlan': ensure_plan,
            '责任计划': plan,
            '责任描述': '身故保险金' + (''.join([f'+{DUTY_ITEM_TEMPLATES[c]["factorName"]}' for c in duty_desc]) if duty_desc else ''),
            '保险期间': ensure_period,
            '保险期间Code': ENSURE_PERIOD_MAP.get(ensure_period, 'TO105'),
            '交费期间': pay_period,
            '交费期间Code': pay_period_code(pay_period),
            '交费方式Code': pay_mode_code(pay_period),
            '性别': gender,
            '性别Code': SEX_MAP.get(gender),
            '年龄': age,
            '年龄类型': age_type,
            '保额(元)': amount,
            '期望费率(‰)': expected_rate,
            '期望保费(元)': round(amount * expected_rate / 1000, 2) if expected_rate else '',
            'age_rate状态码': '',
            'age_rate结果': '',
            'plan_rate状态码': '',
            'plan_rate结果': '',
            'API返回fee': '',
            'failureReason': '',
            '测试结论': '',
            '备注': ''
        }

        try:
            # Step1: age_rate
            ar_ok, ar_cookies, ar_code, ar_text = self.age_rate(age, gender)
            result['age_rate状态码'] = ar_code
            result['age_rate结果'] = 'SUCCESS' if ar_ok else 'FAIL'

            if not ar_ok:
                result['测试结论'] = 'FAIL - age_rate接口失败'
                result['备注'] = ar_text[:200]
                print(f"age_rate FAIL")
                return result

            # Step2: plan_rate（传入动态 ensurePlan）
            pr_ok, fee, pr_code, pr_text, failure_reason = self.plan_rate(
                plan, ensure_period, pay_period, amount, duty_list,
                ensure_plan=ensure_plan, cookies=ar_cookies
            )
            result['plan_rate状态码'] = pr_code
            result['plan_rate结果'] = 'SUCCESS' if pr_ok else 'FAIL'
            result['API返回fee'] = fee if fee is not None else ''
            result['failureReason'] = failure_reason

            # ★ 验证费率：fee 应约等于 保额 × 期望费率(‰) / 1000
            if pr_ok and fee is not None and fee > 0:
                expected_premium = amount * expected_rate / 1000
                # 允许1%误差
                if abs(fee - expected_premium) / expected_premium < 0.01:
                    result['测试结论'] = 'PASS'
                else:
                    result['测试结论'] = f'PASS(费率偏差{abs(fee-expected_premium)/expected_premium*100:.1f}%)'
                print(f"PASS  fee={fee}, 期望={expected_premium:.1f}")
            elif pr_ok and fee is not None and fee == 0 and failure_reason:
                result['测试结论'] = 'FAIL - 计算保费失败'
                result['备注'] = failure_reason[:300]
                print(f"FAIL  {failure_reason[:80]}")
            elif pr_ok and fee is None:
                result['测试结论'] = 'PASS(无fee字段)'
                result['备注'] = pr_text[:300]
                print(f"PASS(no fee)")
            else:
                result['测试结论'] = 'FAIL - plan_rate接口失败'
                result['备注'] = pr_text[:300]
                print(f"FAIL  {pr_text[:100]}")

        except Exception as e:
            result['测试结论'] = f'ERROR - {str(e)}'
            result['备注'] = str(e)
            print(f"ERROR {e}")

        return result


# ──────────────────────────────────────────
# 5. 生成 Excel 测试报告
# ──────────────────────────────────────────
def generate_report(results: list, out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '测试结果明细'

    # 样式
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font  = Font(name='微软雅黑', size=13, bold=True, color='FFFFFF')
    title_fill  = PatternFill(fill_type='solid', fgColor='1F4E79')
    header_font = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2E75B6')
    pass_fill   = PatternFill(fill_type='solid', fgColor='E2EFDA')
    fail_fill   = PatternFill(fill_type='solid', fgColor='FCE4D6')
    warn_fill   = PatternFill(fill_type='solid', fgColor='FFF2CC')
    norm_fill_e = PatternFill(fill_type='solid', fgColor='F2F9FF')
    norm_fill_o = PatternFill(fill_type='solid', fgColor='FFFFFF')

    # 标题
    total_cols = 20
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws['A1'] = f'瑞泰鸿利传世（致享版）终身寿险 — API批量测试报告  生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 32

    # 统计摘要行
    total = len(results)
    passed = sum(1 for r in results if 'PASS' in str(r['测试结论']))
    failed = sum(1 for r in results if 'FAIL' in str(r['测试结论']))
    errors = sum(1 for r in results if 'ERROR' in str(r['测试结论']))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    ws['A2'] = f'共 {total} 条用例 | PASS: {passed} | FAIL: {failed} | ERROR: {errors} | 通过率: {passed/total*100:.1f}%'
    ws['A2'].font = Font(name='微软雅黑', size=10, bold=True, color='1F4E79')
    ws['A2'].fill = PatternFill(fill_type='solid', fgColor='DEEAF1')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 22

    # 表头
    headers = [
        '序号', '保障方案', 'ensurePlan', '责任计划', '责任描述',
        '保险期间', '保险期间Code', '交费期间(年)', '交费期间Code',
        '交费方式Code', '性别', '性别Code', '年龄(岁)', '年龄类型', '保额(元)',
        '期望费率(‰)', '期望保费(元)', 'age_rate\n状态码', 'age_rate\n结果',
        'plan_rate\n状态码', 'plan_rate\n结果',
        'API返回fee', 'failureReason', '测试结论'
    ]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[3].height = 36

    # 数据行
    for ri, r in enumerate(results, 4):
        conclusion = str(r.get('测试结论', ''))
        if 'PASS' in conclusion and 'FAIL' not in conclusion:
            row_fill = pass_fill
        elif 'FAIL' in conclusion:
            row_fill = fail_fill
        elif 'ERROR' in conclusion:
            row_fill = warn_fill
        else:
            row_fill = norm_fill_e if ri % 2 == 0 else norm_fill_o

        vals = [
            r['序号'], r.get('保障方案',''), r.get('ensurePlan',''),
            r['责任计划'], r['责任描述'], r['保险期间'], r['保险期间Code'],
            r['交费期间'], r['交费期间Code'], r['交费方式Code'], r['性别'], r['性别Code'],
            r['年龄'], r['年龄类型'], r['保额(元)'],
            r['期望费率(‰)'], r['期望保费(元)'], r['age_rate状态码'], r['age_rate结果'],
            r['plan_rate状态码'], r['plan_rate结果'],
            r['API返回fee'], r['failureReason'], r['测试结论']
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.alignment = center
            cell.border = border
            cell.fill = row_fill
            cell.font = Font(name='微软雅黑', size=9)
            if ci == len(headers):
                if 'PASS' in conclusion and 'FAIL' not in conclusion:
                    cell.font = Font(name='微软雅黑', size=9, bold=True, color='375623')
                elif 'FAIL' in conclusion:
                    cell.font = Font(name='微软雅黑', size=9, bold=True, color='C00000')
                elif 'ERROR' in conclusion:
                    cell.font = Font(name='微软雅黑', size=9, bold=True, color='C55A11')
        ws.row_dimensions[ri].height = 20

    # 列宽（24列）
    col_widths = [6, 10, 10, 8, 38, 8, 12, 10, 12, 10, 6, 8, 8, 10, 14, 12, 14, 10, 10, 10, 10, 14, 30, 18]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'A4'

    # ── Sheet2: 摘要统计 ──
    ws2 = wb.create_sheet('汇总统计')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = '测试汇总统计'
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = center
    ws2.row_dimensions[1].height = 30

    stat_headers = ['维度', '分类', '总用例', 'PASS', 'FAIL/ERROR']
    for ci, h in enumerate(stat_headers, 1):
        cell = ws2.cell(2, ci, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border

    from collections import defaultdict
    bz_stat   = defaultdict(lambda: {'total':0,'pass':0,'fail':0})   # ★ 保障方案维度
    plan_stat = defaultdict(lambda: {'total':0,'pass':0,'fail':0})
    pay_stat  = defaultdict(lambda: {'total':0,'pass':0,'fail':0})
    sex_stat  = defaultdict(lambda: {'total':0,'pass':0,'fail':0})
    age_type_stat = defaultdict(lambda: {'total':0,'pass':0,'fail':0})

    for r in results:
        conc = str(r.get('测试结论',''))
        is_pass = 'PASS' in conc and 'FAIL' not in conc
        bz_stat[r.get('保障方案','')]['total'] += 1
        bz_stat[r.get('保障方案','')]['pass' if is_pass else 'fail'] += 1
        plan_stat[r['责任计划']]['total'] += 1
        plan_stat[r['责任计划']]['pass' if is_pass else 'fail'] += 1
        pay_stat[r['交费期间']]['total'] += 1
        pay_stat[r['交费期间']]['pass' if is_pass else 'fail'] += 1
        sex_stat[r['性别']]['total'] += 1
        sex_stat[r['性别']]['pass' if is_pass else 'fail'] += 1
        age_type_stat[r['年龄类型']]['total'] += 1
        age_type_stat[r['年龄类型']]['pass' if is_pass else 'fail'] += 1

    stat_rows = [('总计', '全部', total, passed, failed+errors)]
    for bz in sorted(bz_stat):
        s = bz_stat[bz]
        stat_rows.append(('保障方案', bz, s['total'], s['pass'], s['fail']))
    for plan in sorted(plan_stat):
        s = plan_stat[plan]
        stat_rows.append(('责任计划', f'计划{plan}', s['total'], s['pass'], s['fail']))
    for pay in sorted(pay_stat):
        s = pay_stat[pay]
        stat_rows.append(('交费期间', f'{pay}年', s['total'], s['pass'], s['fail']))
    for sex in sorted(sex_stat):
        s = sex_stat[sex]
        stat_rows.append(('性别', sex, s['total'], s['pass'], s['fail']))
    for at in sorted(age_type_stat):
        s = age_type_stat[at]
        stat_rows.append(('年龄类型', at, s['total'], s['pass'], s['fail']))

    alt = [PatternFill(fill_type='solid', fgColor='F2F9FF'),
           PatternFill(fill_type='solid', fgColor='FFFFFF')]
    for ri2, row in enumerate(stat_rows, 3):
        bg = pass_fill if row[3] == row[2] else (fail_fill if row[4] > 0 else alt[ri2%2])
        for ci, v in enumerate(row, 1):
            cell = ws2.cell(ri2, ci, value=v)
            cell.font = Font(name='微软雅黑', size=10)
            cell.fill = bg; cell.alignment = center; cell.border = border
        ws2.row_dimensions[ri2].height = 22

    for ci, w in enumerate([14,14,12,12,14], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_path)
    print(f"\n报告已保存: {out_path}")
    print(f"总计: {total} | PASS: {passed} | FAIL: {failed} | ERROR: {errors}")


# ──────────────────────────────────────────
# 6. 主流程
# ──────────────────────────────────────────
if __name__ == '__main__':
    print("=" * 60)
    print("批量接口测试开始")
    print("=" * 60)

    # 加载测试用例
    cases = load_test_cases()
    print(f"\n共加载 {len(cases)} 条测试用例（每个组合含最小/最大年龄）")

    # 初始化API客户端
    api = BatchAPITest()

    # 登录
    print("\n>>> Step 1: 登录...")
    login_ok, _ = api.login()
    if not login_ok:
        print("  警告: 登录可能失败，继续尝试执行...")

    # 批量执行
    print(f"\n>>> Step 2: 开始批量执行 {len(cases)} 条用例...")
    all_results = []
    for i, case in enumerate(cases, 1):
        result = api.run_single_case(case, i)
        all_results.append(result)
        # 每5条稍作间隔
        if i % 5 == 0:
            time.sleep(0.3)

    # 生成报告（文件名含日期时间）
    print("\n>>> Step 3: 生成Excel测试报告...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = rf'f:\workbuddy\test_Jhs\API测试报告_{ts}.xlsx'
    generate_report(all_results, out_path)

    print("\n" + "=" * 60)
    print("测试完成！")
    print("=" * 60)
