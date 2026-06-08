
import openpyxl
import json

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
ws = wb['标准体费率表']

# 读取表头行
row4 = {c: ws.cell(row=4, column=c).value for c in range(2, 114)}  # 责任计划
row5 = {c: ws.cell(row=5, column=c).value for c in range(2, 114)}  # 保险期间
row6 = {c: ws.cell(row=6, column=c).value for c in range(2, 114)}  # 交费期间
row7 = {c: str(ws.cell(row=7, column=c).value).strip() for c in range(2, 114)}  # 性别

# 找年龄的最小行和最大行（8~77行，年龄在第1列）
age_rows = {}
for r in range(8, 78):
    age = ws.cell(row=r, column=1).value
    if age is not None:
        age_rows[age] = r

min_age = min(age_rows.keys())
max_age = max(age_rows.keys())
print(f"年龄范围: {min_age} ~ {max_age}")
print(f"边界年龄: {min_age}岁 (行{age_rows[min_age]}), {max_age}岁 (行{age_rows[max_age]})")

# 建立列到(计划,保险期间,交费期间,性别)的映射
col_map = {}
for c in range(2, 114):
    plan = row4.get(c)
    period = row5.get(c)
    pay = row6.get(c)
    gender = row7.get(c, '').strip()
    if plan is not None and period and pay and gender in ('男', '女'):
        col_map[c] = {'计划': int(plan), '保险期间': str(period), '交费期间': int(pay), '性别': gender}

# 提取所有(计划,保险期间,交费期间,性别)组合
combos = {}
for c, info in col_map.items():
    key = (info['计划'], info['保险期间'], info['交费期间'], info['性别'])
    if key not in combos:
        combos[key] = []
    combos[key].append(c)

print(f"\n共 {len(combos)} 种组合")
print("\n各组合下列索引（应该各只有1个）:")
multi = {k: v for k, v in combos.items() if len(v) != 1}
if multi:
    print("  存在多列映射到同一组合:")
    for k, v in multi.items():
        print(f"  {k}: 列{v}")
else:
    print("  每种组合恰好对应1列，结构正确")

# 提取边界值
results = []
for key, cols in sorted(combos.items()):
    plan, ins_period, pay_period, gender = key
    col = cols[0]
    min_val = ws.cell(row=age_rows[min_age], column=col).value
    max_val = ws.cell(row=age_rows[max_age], column=col).value
    results.append({
        '责任计划': plan,
        '保险期间': ins_period,
        '交费期间': pay_period,
        '性别': gender,
        f'{min_age}岁费率': min_val,
        f'{max_age}岁费率': max_val
    })

# 输出结果
print(f"\n=== 边界年龄费率提取结果（{min_age}岁 和 {max_age}岁） ===")
print(f"{'责任计划':>6} {'保险期间':>6} {'交费期间':>6} {'性别':>4} {str(min_age)+'岁费率':>10} {str(max_age)+'岁费率':>10}")
print("-" * 60)
for r in results:
    print(f"{r['责任计划']:>6} {r['保险期间']:>6} {r['交费期间']:>6} {r['性别']:>4} {str(r[f'{min_age}岁费率']):>10} {str(r[f'{max_age}岁费率']):>10}")

# 保存为JSON供后续生成Excel
with open(r'f:\workbuddy\test_Jhs\boundary_rates.json', 'w', encoding='utf-8') as f:
    json.dump({'min_age': min_age, 'max_age': max_age, 'results': results}, f, ensure_ascii=False, indent=2)
print("\n数据已保存到 boundary_rates.json")
