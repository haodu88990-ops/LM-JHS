
import openpyxl

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
ws = wb['字段说明']
print(f"字段说明 Sheet: max_row={ws.max_row}, max_col={ws.max_column}")
print("\n全部内容:")
for r in range(1, ws.max_row+1):
    row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
    non_none = [v for v in row if v is not None and v != '']
    if non_none:
        print(f"  行{r}: {row}")
