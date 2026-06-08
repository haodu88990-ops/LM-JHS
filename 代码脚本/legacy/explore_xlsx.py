
import openpyxl

wb = openpyxl.load_workbook(r'f:\workbuddy\test_Jhs\费率表.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===")
    # 打印前30行，每行前15列
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=15, values_only=True):
        print(row)
